# -*- coding: utf-8 -*-

from PyQt4.QtGui import *
from PyQt4.QtCore import *
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from datetime import datetime, date
import time
import pyodbc
import os




# ---------------------------------- PARÁMETROS -------------------------------------------------------------
#       Modificar estos valores de ser necesario:
iva = 0.21
min = 35.56

# ---------------------------------- PARÁMETROS -------------------------------------------------------------

try:
    _fromUtf8 = QString.fromUtf8
except AttributeError:
    def _fromUtf8(s):
        return s

class clase(QWidget):
    def __init__(self, row, tabla, cab):
        super(clase, self).__init__(parent=None)
        self.setObjectName("clase")
        self.lyt = QHBoxLayout()
        self.cb = QComboBox()
        self.cb.setObjectName("cb")
        self.tabla = tabla
        self.cab = cab
        self.row = row
        self.cb.addItems(['LIQ', 'NLIQ', 'INC'])
        self.lyt.setAlignment(Qt.AlignHCenter)
        self.lyt.addWidget(self.cb)
        self.lyt.setContentsMargins(5, 0, 5, 0)
        self.setLayout(self.lyt)
        self.cb.currentIndexChanged.connect(self.emite)
        self.cb.setStyleSheet(_fromUtf8("QComboBox {\n"
                                                      "    border: 2px solid rgb(143,212,0);\n"
                                                      "    border-radius: 5px;\n"
                                                      "    padding: 1px 18px 1px 3px;\n"
                                                      "    min-width: 1em;\n"
                                                      "    color: dark grey;\n"
                                                      "    font-weight: bold;\n"
                                                      "}\n"
                                                      "QComboBox:editable {\n"
                                                      "    background: darkgrey;\n"
                                                      "}\n"
                                                      "\n"
                                                      "QComboBox:!editable, QComboBox::drop-down:editable {\n"
                                                      "     background: white;\n"
                                                      "    \n"
                                                      "}\n"
                                                      "\n"
                                                      "QComboBox:on { /* shift the text when the popup opens */\n"
                                                      "    background: darkgrey;\n"
                                                      "    padding-top: 3px;\n"
                                                      "    padding-left: 4px;\n"
                                                      "}\n"
                                                      "\n"
                                                      "QComboBox::drop-down {\n"
                                                      "    font-weight: light;\n"
                                                      "    subcontrol-origin: padding;\n"
                                                      "    subcontrol-position: top right;\n"
                                                      "    width: 15px;\n"
                                                      "\n"
                                                      "    border-left-width: 1px;\n"
                                                      "    border-left-color: rgb(143,212,0);\n"
                                                      "    border-left-style: solid; /* just a single line */\n"
                                                      "    border-top-right-radius: 3px; /* same radius as the QComboBox */\n"
                                                      "    border-bottom-right-radius: 3px;\n"
                                                      "}\n"
                                                      "\n"
                                                      "QComboBox::down-arrow {\n"
                                                      "    image: url(:/img/desplegado.png);\n"
                                                      "}"))
    def emite(self):
        colorea(self.tabla, self.row, self.cab, self.idx())
    def idx(self):
        return self.cb.currentIndex()



def confTabla(form):
    form.tabla_2.clear()
    form.tabla_2.setRowCount(0)
    cabeceras = [u'CLASE', u'COD MUNI', u'MUNICIPIO', u'REF CATASTRAL', u'NUM FIJO', u'OBJETO TRIBUTARIO', u'ID_VALOR',
                 u'OTROS VALORES', u'PERIODO', u'VALOR CAT', u'BASE LIQ', u'Tipo', u'DEUDA LIQ', u'IMP PADRON 2016',
                 u'17.185% CUOTA', u'FACTURACION CON IVA', u'FACTURACION SIN IVA', u'SUJETO PASIVO', u'NIF',
                 u'Nº EXP GERENCIA', u'TIPO EXP', u'OBSERVACIONES']


    form.tabla_2.setColumnCount(len(cabeceras))
    for i in range(form.tabla.rowCount()):
        if form.tabla.cellWidget(i, 0).isChecked():
            refcat = form.tabla.item(i, 12).text()
            expger = form.tabla.item(i, 9).text()
            expgtt = form.tabla.item(i, 7).text()
            municipio = form.tabla.item(i, 5).text()
            id_subo = form.tabla.item(i, 4).text()
            if form.tabla.item(i, 11).text() == u'GRÁFICO':
                obs = u'AJUSTE GRÁFICO'
            else:
                obs = ""
            objtributario = form.tabla.item(i, 15).text()
            lcargos = DevuelveCargos(form.tabla.item(i, 17).text())
            falteracion = datetime.strptime(str(form.tabla.item(i, 14).text()), '%Y-%m-%d %H:%M:%S').__format__(
                '%Y')
            if form.tabla.item(i, 21).text() != "None":
                fentrega = datetime.strptime(str(form.tabla.item(i, 21).text()), '%Y-%m-%d %H:%M:%S').__format__(
                    '%Y')
            else:
                fentrega = ""
            for i in range(0, len(lcargos), 1):
                QApplication.processEvents()
                row = form.tabla_2.rowCount()
                form.tabla_2.setRowCount(form.tabla_2.rowCount() + 1)
                cl = clase(row, form.tabla_2, len(cabeceras))
                cl.setObjectName("cl")
                if fentrega != "":
                    if int(fentrega) == int(falteracion):
                        cl.cb.setCurrentIndex(2)
                    elif int(fentrega) > int(falteracion):
                        cl.cb.setCurrentIndex(1)
                form.tabla_2.setCellWidget(row, 0, cl)
                form.tabla_2.setItem(row, 1, QTableWidgetItem(id_subo[-3:]))
                form.tabla_2.setItem(row, 2, QTableWidgetItem(municipio))
                form.tabla_2.setItem(row, 3, QTableWidgetItem(refcat + componeCargo(lcargos[i])))
                form.tabla_2.setItem(row, 5, QTableWidgetItem(objtributario))
                form.tabla_2.setItem(row, 8, QTableWidgetItem(str(date.today().year)))
                form.tabla_2.setItem(row, 11, QTableWidgetItem(getTI(id_subo)))
                form.tabla_2.setItem(row, 19, QTableWidgetItem(expger))
                form.tabla_2.setItem(row, 21, QTableWidgetItem(obs))
                colorea(form.tabla_2, row, len(cabeceras), cl.cb.currentIndex())

    form.tabla_2.setHorizontalHeaderLabels(cabeceras)
    form.tabla_2.horizontalHeader().setResizeMode(QHeaderView.ResizeToContents)




def componeCargo(cargo):
    cargo = str(cargo)
    while len(cargo) < 4:
        cargo = "0" + cargo
    return cargo

def colorea(tabla, row, cab, idx):
    if idx == 2:
        color = QColor(qRgb(255, 250, 205))
    elif idx == 1:
        color = QColor(qRgb(194, 255, 220))
    elif idx == 0:
        color = QColor(qRgb(173, 216, 230))
    for a in range(cab):
        if tabla.item(row, a) is not None:
            tabla.item(row, a).setBackground(color)
        else:
            tabla.setItem(row, a, QTableWidgetItem(""))
            tabla.item(row, a).setBackground(color)




def activarFunciones(form):
    form.desactivarFunciones()
    form.actionImportar_datos_desde_FIN_de_Retorno.setEnabled(True)
    form.actionImportar_datos_desde_FIN_de_Retorno.setText("Importar FIN")
    form.but_Calcular.setEnabled(True)
    form.actionImportar_Access_Valores.setVisible(True)


def configura(form):
    pass


def calcular(form):
    for i in range(form.tabla_2.rowCount()):
        QApplication.processEvents()
        if form.tabla_2.item(i, 6) is None or (
                        form.tabla_2.item(i, 6) is not None and form.tabla_2.item(i, 6).text() == ""):
            if form.tabla_2.item(i, 11).text() != "MANUAL":
                bliq = float(form.tabla_2.item(i, 10).text())
                tipo = float(form.tabla_2.item(i, 11).text())
                pad = round(bliq * tipo / 100, 2)
                cuota = round(pad * 0.17185, 2)
                if cuota < min:
                    fact = min
                else:
                    fact = cuota
                factiva = round(fact * 1.21, 2)
                form.tabla_2.setItem(i, 13, QTableWidgetItem(str(pad)))
                form.tabla_2.setItem(i, 14, QTableWidgetItem(str(cuota)))
                form.tabla_2.setItem(i, 15, QTableWidgetItem(str(factiva)))
                form.tabla_2.setItem(i, 16, QTableWidgetItem(str(fact)))
                colorea(form.tabla_2, i, len(form.tabla_2.horizontalHeader()), form.tabla_2.cellWidget(i, 0).idx())
        else:
            liq = float(form.tabla_2.item(i, 12).text())
            cuota = round(liq * 0.17185, 2)
            if cuota < min:
                fact = min
            else:
                fact = cuota
            factiva = round(fact * 1.21, 2)
            form.tabla_2.setItem(i, 14, QTableWidgetItem(str(cuota)))
            form.tabla_2.setItem(i, 15, QTableWidgetItem(str(factiva)))
            form.tabla_2.setItem(i, 16, QTableWidgetItem(str(fact)))
            colorea(form.tabla_2, i, len(form.tabla_2.horizontalHeader()), form.tabla_2.cellWidget(i, 0).idx())




def importaFin(form):
    ruta = QFileDialog.getOpenFileName(form, "Seleccionar fichero FIN", "", "FIN RETORNO (*.*)")
    if ruta != "":
        file = open(unicode(ruta), "r")
        p = form.loading("Leyendo FIN", form.tabWidget)
        p.show()
        for line in file:
            QApplication.processEvents()
            if line[0:2] == "01":
                print line[53:57]
                tipofin = line[53:57]
            if line[0:2] == "15":
                obj = line[200:205].strip() + " " + line[205:230].strip() + " " + str(int(line[230:234])) + line[
                                                                                                            234:235]
                refcat = line[30:50]
                expediente = line[527:535] + "." + line[536:538] + "/" + line[523:525]
                tipoexp = line[509:513]
                escalera = line[249:251].strip()
                planta = line[251:254].strip()
                puerta = line[254:257].strip()
                nfijo = line[50:58]
                vcat = str(float(line[379:391]) / 100)
                bliq = str(float(line[415:427]) / 100)
                if escalera != "":
                    escalera = " ES:" + escalera
                if planta != "":
                    planta = " PL:" + planta
                if puerta != "":
                    puerta = " PT:" + puerta
                obj = obj + escalera + planta + puerta
                for i in range(form.tabla_2.rowCount()):
                    if form.tabla_2.item(i, 3).text()[:18] == refcat[:18]:
                        form.tabla_2.setItem(i, 3, QTableWidgetItem(refcat))
                        form.tabla_2.setItem(i, 5, QTableWidgetItem(obj))
                        form.tabla_2.setItem(i, 4, QTableWidgetItem(nfijo))
                        form.tabla_2.setItem(i, 9, QTableWidgetItem(vcat))
                        form.tabla_2.setItem(i, 10, QTableWidgetItem(bliq))
                        if form.tabla_2.item(i, 19).text() == expediente:
                            form.tabla_2.setItem(i, 20, QTableWidgetItem(tipoexp))
                        colorea(form.tabla_2, i, len(form.tabla_2.horizontalHeader()), form.tabla_2.cellWidget(i, 0).idx())

            if line[0:2] == "46":
                if line[57:60] == "001":
                    refcat = line[30:48]
                    titular = line[69:129]
                    titular = titular.strip()
                    nif = line[60:69]
                    for i in range(form.tabla_2.rowCount()):
                        if form.tabla_2.item(i, 3).text()[:18] == refcat:
                            form.tabla_2.setItem(i, 17, QTableWidgetItem(titular))
                            form.tabla_2.setItem(i, 18, QTableWidgetItem(nif))
                            colorea(form.tabla_2, i, len(form.tabla_2.horizontalHeader()),
                                    form.tabla_2.cellWidget(i, 0).idx())
        p.close()


def StringCargos(uu):
    cargos = []
    literalcargos = []
    literalcargos.append(uu.split("|")[0])
    literalcargos.append(uu.split("|")[1])
    for literal in literalcargos:
        previocargos = literal.split(";")
        for a in previocargos:
            if a[:1] == "[":
                intcargos = a[1:-1].split("-")
                for b in range(int(intcargos[0]), int(intcargos[1]) + 1, 1):
                    cargos.append(b)
            else:
                if a != "":
                    cargos.append(a)
    if len(cargos) == 1:
        devolucion = "CARGO "
    else:
        devolucion = "CARGOS "
    for cargo in cargos:
        if cargo != "":
            devolucion = devolucion + str(cargo) + ", "
    return devolucion[:-2]


def DevuelveCargos(uu):
    cargos = []
    literalcargos = []
    literalcargos.append(uu.split("|")[0])
    literalcargos.append(uu.split("|")[1])
    for literal in literalcargos:
        previocargos = literal.split(";")
        for a in previocargos:
            if a[:1] == "[":
                intcargos = a[1:-1].split("-")
                for b in range(int(intcargos[0]), int(intcargos[1]) + 1, 1):
                    cargos.append(b)
            else:
                if a != "":
                    cargos.append(a)
    return cargos


def descomponeUU(uu):
    a = ['', '']
    if uu >= 16:
        a[0] = 16
        a[1] = uu - 16
    elif uu < 16:
        a[0] = uu
        a[1] = 0
    return a


def exportar(form):
    ruta = QFileDialog.getOpenFileName(form, u"Seleccionar fichero de facturación", "", "Ficheros Excel (*.xls *.xlsx)")
    if ruta != "":
        wb = openpyxl.load_workbook(unicode(ruta))
        nhoja, ok = QInputDialog.getText(form, "Nombre de la nueva hoja para LIQUIDADAS", "Introduzca nombre:")
        hj = wb.create_sheet(str(nhoja) + " LIQ")
        bg = PatternFill(fill_type='solid', start_color='FF8C00', end_color='FF8C00')
        ag = Alignment(horizontal='center', vertical='center')
        borde = Border(left=Side(border_style='thin', color='000000'),
                       right=Side(border_style='thin', color='000000'),
                       top=Side(border_style='thin', color='000000'),
                       bottom=Side(border_style='thin', color='000000'))
        fuente = Font(name='Calibri',
                      size=10,
                      bold=True,
                      color='000000')

        for c in range(1, form.tabla_2.columnCount(), 1):
            d = hj.cell(row=1, column=c, value=str(form.tabla_2.horizontalHeaderItem(c).text()))
            d.fill = bg
            d.alignment = ag
            d.border = borde
            d.font = fuente
        anchos = [('A', 9.28515625), ('B', 25.85546875), ('C', 23.28515625), ('D', 9.85546875), ('E', 40.42578125),
                  ('F', 23.28515625), ('G', 13.85546875), ('H', 11.0), ('I', 11.28515625), ('J', 11.5703125),
                  ('K', 9.42578125), ('L', 9.42578125), ('M', 15.7109375), ('N', 13.5703125), ('O', 19.42578125),
                  ('P', 18.42578125), ('Q', 46.0), ('R', 11.140625), ('S', 17.5703125), ('T', 8.85546875), ('U', 20.85546875)]



        for columna, ancho in anchos:
            hj.column_dimensions[columna].width = ancho
        borde = Border(left=Side(border_style='thin', color='00000000'),
                       right=Side(border_style='thin', color='00000000'),
                       top=Side(border_style='thin', color='00000000'),
                       bottom=Side(border_style='thin', color='00000000'))
        fuente = Font(name='Calibri', size=10)
        contador = 1
        for fila in range(form.tabla_2.rowCount()):
            if form.tabla_2.cellWidget(fila, 0).idx() == 0:
                contador += 1

                fondoverde = PatternFill(fill_type='solid', start_color='E2EFDA', end_color='E2EFDA')
                for columna in range(1, form.tabla_2.columnCount()):
                    if form.tabla_2.item(fila, columna) is not None:
                        print contador, columna, str(form.tabla_2.item(fila, columna).text())
                        d = hj.cell(row=contador, column=columna, value=str(form.tabla_2.item(fila, columna).text()))

                    else:
                        d = hj.cell(row=contador, column=columna, value="")
                    d.border = borde
                    d.font = fuente


        for col in hj.columns:
            if col[0].column in ['I', 'J', 'L', 'M', 'N', 'O', 'P']:
                for cell in col:
                    try:
                        a = float(cell.value)
                        cell.value = a
                        cell.number_format = "#,##0.00"
                    except:
                        pass
            elif col[0].column in ['K']:
                for cell in col:
                    try:
                        a = float(cell.value)
                        cell.value = a
                        cell.number_format = "0.0000"
                    except:
                        pass
        totalfilas = contador
        borde = Border(left=Side(border_style='thin', color='00000000'),
                       right=Side(border_style='thin', color='00000000'),
                       top=Side(border_style='thin', color='00000000'),
                       bottom=Side(border_style='thin', color='00000000'))
        fuente = Font(name='Calibri',
                      size=10,
                      bold=True)
        bg = PatternFill(fill_type='solid', start_color='FF8C00', end_color='FF8C00')
        a = hj.cell(row=totalfilas + 1, column=15, value="CON IVA")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a.alignment = Alignment(horizontal="center")
        a = hj.cell(row=totalfilas + 1, column=16, value="SIN IVA")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a.alignment = Alignment(horizontal="center")
        a = hj.cell(row=totalfilas + 2, column=17, value="TOTAL FACTURADO LIQUIDADAS")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a.alignment = Alignment(horizontal="center")
        a = hj.cell(row=totalfilas + 2, column=15, value="=SUM(O2:O" + str(totalfilas) + ")")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a.alignment = Alignment(horizontal="center")
        a.number_format = "#,##0.00"
        a = hj.cell(row=totalfilas + 2, column=16, value="=SUM(P2:P" + str(totalfilas) + ")")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a.alignment = Alignment(horizontal="center")
        a.number_format = "#,##0.00"

        # ------------------------------------NO LIQUIDADAS-------------------------------------------------------------

        hj = wb.create_sheet(str(nhoja) + " NLQ")
        bg = PatternFill(fill_type='solid', start_color='FF8C00', end_color='FF8C00')
        ag = Alignment(horizontal='center', vertical='center')
        borde = Border(left=Side(border_style='thin', color='000000'),
                       right=Side(border_style='thin', color='000000'),
                       top=Side(border_style='thin', color='000000'),
                       bottom=Side(border_style='thin', color='000000'))
        fuente = Font(name='Calibri',
                      size=10,
                      bold=True,
                      color='000000')

        for c in range(1, form.tabla_2.columnCount(), 1):
            d = hj.cell(row=1, column=c, value=str(form.tabla_2.horizontalHeaderItem(c).text()))
            d.fill = bg
            d.alignment = ag
            d.border = borde
            d.font = fuente
        anchos = [('A', 9.28515625), ('B', 25.85546875), ('C', 23.28515625), ('D', 9.85546875), ('E', 40.42578125),
                  ('F', 23.28515625), ('G', 13.85546875), ('H', 11.0), ('I', 11.28515625), ('J', 11.5703125),
                  ('K', 9.42578125), ('L', 9.42578125), ('M', 15.7109375), ('N', 13.5703125), ('O', 19.42578125),
                  ('P', 18.42578125), ('Q', 46.0), ('R', 11.140625), ('S', 17.5703125), ('T', 8.85546875),
                  ('U', 20.85546875)]

        for columna, ancho in anchos:
            hj.column_dimensions[columna].width = ancho
        borde = Border(left=Side(border_style='thin', color='00000000'),
                       right=Side(border_style='thin', color='00000000'),
                       top=Side(border_style='thin', color='00000000'),
                       bottom=Side(border_style='thin', color='00000000'))
        fuente = Font(name='Calibri', size=10)
        contador = 1
        for fila in range(form.tabla_2.rowCount()):
            if form.tabla_2.cellWidget(fila, 0).idx() == 1:
                contador += 1

                fondoverde = PatternFill(fill_type='solid', start_color='E2EFDA', end_color='E2EFDA')
                for columna in range(1, form.tabla_2.columnCount()):
                    if form.tabla_2.item(fila, columna) is not None:
                        print contador, columna, str(form.tabla_2.item(fila, columna).text())
                        d = hj.cell(row=contador, column=columna, value=str(form.tabla_2.item(fila, columna).text()))

                    else:
                        d = hj.cell(row=contador, column=columna, value="")
                    d.border = borde
                    d.font = fuente

        for col in hj.columns:
            if col[0].column in ['I', 'J', 'L', 'M', 'N', 'O', 'P']:
                for cell in col:
                    try:
                        a = float(cell.value)
                        cell.value = a
                        cell.number_format = "#,##0.00"
                    except:
                        pass
            elif col[0].column in ['K']:
                for cell in col:
                    try:
                        a = float(cell.value)
                        cell.value = a
                        cell.number_format = "0.0000"
                    except:
                        pass
        totalfilas = contador
        borde = Border(left=Side(border_style='thin', color='00000000'),
                       right=Side(border_style='thin', color='00000000'),
                       top=Side(border_style='thin', color='00000000'),
                       bottom=Side(border_style='thin', color='00000000'))
        fuente = Font(name='Calibri',
                      size=10,
                      bold=True)
        bg = PatternFill(fill_type='solid', start_color='FF8C00', end_color='FF8C00')
        a = hj.cell(row=totalfilas + 1, column=15, value="CON IVA")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a.alignment = Alignment(horizontal="center")
        a = hj.cell(row=totalfilas + 1, column=16, value="SIN IVA")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a.alignment = Alignment(horizontal="center")
        a = hj.cell(row=totalfilas + 2, column=17, value="TOTAL FACTURADO NO LIQUIDADAS")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a.alignment = Alignment(horizontal="center")
        a = hj.cell(row=totalfilas + 2, column=15, value="=SUM(O2:O" + str(totalfilas) + ")")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a.alignment = Alignment(horizontal="center")
        a.number_format = "#,##0.00"
        a = hj.cell(row=totalfilas + 2, column=16, value="=SUM(P2:P" + str(totalfilas) + ")")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a.alignment = Alignment(horizontal="center")
        a.number_format = "#,##0.00"

        #------------------------------------- INC. PADRON -------------------------------------------------------------

        bg = PatternFill(fill_type='solid', start_color='3CB371', end_color='3CB371')
        contador = totalfilas + 5
        iniciocuenta = totalfilas + 6
        for c in range(1, form.tabla_2.columnCount(), 1):
            d = hj.cell(row=contador, column=c, value=str(form.tabla_2.horizontalHeaderItem(c).text()))
            d.fill = bg
            d.alignment = ag
            d.border = borde
            d.font = fuente
        anchos = [('A', 9.28515625), ('B', 25.85546875), ('C', 23.28515625), ('D', 9.85546875), ('E', 40.42578125),
                  ('F', 23.28515625), ('G', 13.85546875), ('H', 11.0), ('I', 11.28515625), ('J', 11.5703125),
                  ('K', 9.42578125), ('L', 9.42578125), ('M', 15.7109375), ('N', 13.5703125), ('O', 19.42578125),
                  ('P', 18.42578125), ('Q', 46.0), ('R', 11.140625), ('S', 17.5703125), ('T', 8.85546875),
                  ('U', 20.85546875)]

        for columna, ancho in anchos:
            hj.column_dimensions[columna].width = ancho
        borde = Border(left=Side(border_style='thin', color='00000000'),
                       right=Side(border_style='thin', color='00000000'),
                       top=Side(border_style='thin', color='00000000'),
                       bottom=Side(border_style='thin', color='00000000'))
        fuente = Font(name='Calibri', size=10)

        for fila in range(form.tabla_2.rowCount()):
            if form.tabla_2.cellWidget(fila, 0).idx() == 2:
                contador += 1

                fondoverde = PatternFill(fill_type='solid', start_color='E2EFDA', end_color='E2EFDA')
                for columna in range(1, form.tabla_2.columnCount()):
                    if form.tabla_2.item(fila, columna) is not None:
                        print contador, columna, str(form.tabla_2.item(fila, columna).text())
                        d = hj.cell(row=contador, column=columna, value=str(form.tabla_2.item(fila, columna).text()))

                    else:
                        d = hj.cell(row=contador, column=columna, value="")
                    d.border = borde
                    d.font = fuente

        for col in hj.columns:
            if col[0].column in ['I', 'J', 'L', 'M', 'N', 'O', 'P']:
                for cell in col:
                    try:
                        a = float(cell.value)
                        cell.value = a
                        cell.number_format = "#,##0.00"
                    except:
                        pass
            elif col[0].column in ['K']:
                for cell in col:
                    try:
                        a = float(cell.value)
                        cell.value = a
                        cell.number_format = "0.0000"
                    except:
                        pass
        totalfilas = contador
        borde = Border(left=Side(border_style='thin', color='00000000'),
                       right=Side(border_style='thin', color='00000000'),
                       top=Side(border_style='thin', color='00000000'),
                       bottom=Side(border_style='thin', color='00000000'))
        fuente = Font(name='Calibri',
                      size=10,
                      bold=True)
        bg = PatternFill(fill_type='solid', start_color='FF8C00', end_color='FF8C00')
        a = hj.cell(row=totalfilas + 1, column=15, value="CON IVA")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a.alignment = Alignment(horizontal="center")
        a = hj.cell(row=totalfilas + 1, column=16, value="SIN IVA")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a.alignment = Alignment(horizontal="center")
        a = hj.cell(row=totalfilas + 2, column=17, value="TOTAL FACTURADO INCORPORACIÓN PADRÓN")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a.alignment = Alignment(horizontal="center")
        a = hj.cell(row=totalfilas + 2, column=15, value="=SUM(O" + str(iniciocuenta) + ":O" + str(totalfilas) + ")")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a.alignment = Alignment(horizontal="center")
        a.number_format = "#,##0.00"
        a = hj.cell(row=totalfilas + 2, column=16, value="=SUM(P" + str(iniciocuenta) + ":P" + str(totalfilas) + ")")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a.alignment = Alignment(horizontal="center")
        a.number_format = "#,##0.00"

        try:
            wb.save(unicode(ruta))
            QMessageBox.information(form, "Guardado", u"Guardado correctamente")
        except IOError as e:
            if e.errno == 13:
                QMessageBox.critical(form, "Error", u"El fichero de destino está en uso")


def getTI(id_subo):
    path = os.getcwd() + "/profiles/tipos.accdb"

    conn_str = (
        r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
        r'DBQ=' + path + ';'
    )
    cnxn = pyodbc.connect(conn_str)
    crsr = cnxn.cursor()
    sql = "SELECT TIPO_IMPOSITIVO FROM TI WHERE ID_SUBO=" + str(id_subo)
    crsr.execute(sql)
    row = crsr.fetchone()
    if row:
        return str(row.TIPO_IMPOSITIVO).replace(",", ".")
    else:
        return "MANUAL"


def CargaValores(form):
    path = QFileDialog.getOpenFileName(form, u"Seleccionar Access de Explotación", "", "Ficheros Excel (*.mdb *.accdb)")
    path = unicode(path)
    p = form.loading("Importando valores desde Access", form.tabWidget)
    p.show()
    conn_str = (
        r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
        r'DBQ=' + path + ';'
    )
    cnxn = pyodbc.connect(conn_str)
    crsr = cnxn.cursor()
    sql = "SELECT temp.REFERENCIA_CATASTRAL AS RC, NIF_SP_VALO AS NIF, temp.ID_VALO, temp.VALORES, NOMBRE_SP_VALO AS SP, temp.IMPORTE AS IMP, " \
          "SWITCH(temp.PER = temp.MINPER, temp.MINPER," \
          "temp.PER <> temp.MINPER, temp.MINPER & '-' & temp.PER) AS PERI " \
          "FROM(" \
          "SELECT NOMBRE_SP_VALO, NIF_SP_VALO, REFERENCIA_CATASTRAL, " \
          "CINT(PERIODO) AS PER FROM [LISTADO 1] ) AS ini " \
          "INNER JOIN (" \
          "SELECT MAX(CINT(PERIODO))AS PER, SUM(IMPORTE_VALO) AS IMPORTE, " \
          "MIN(CINT(PERIODO)) AS MINPER, REFERENCIA_CATASTRAL, FIRST(ID_VALOR) AS ID_VALO, COUNT(ID_VALOR) AS VALORES" \
          " FROM [LISTADO 1] GROUP BY REFERENCIA_CATASTRAL) as temp " \
          "ON (ini.REFERENCIA_CATASTRAL = temp.REFERENCIA_CATASTRAL AND ini.PER= temp.PER);"

    crsr.execute(sql)
    rows = crsr.fetchall()
    for row in rows:
        RC = row.RC
        SP = row.SP
        NIF = row.NIF
        PERI = row.PERI
        IMP = row.IMP
        IMP = float(IMP) / 100
        ID = row.ID_VALO
        if row.VALORES > 1:
            CIND = int(row.VALORES) - 1
        else:
            CIND = ""

        for i in range(form.tabla_2.rowCount()):
            QApplication.processEvents()
            refcat = form.tabla_2.item(i, 3).text()
            if RC == refcat:
                form.tabla_2.setItem(i, 17, QTableWidgetItem(str(SP)))
                form.tabla_2.setItem(i, 18, QTableWidgetItem(str(NIF)))
                form.tabla_2.setItem(i, 12, QTableWidgetItem(str(IMP)))
                form.tabla_2.setItem(i, 8, QTableWidgetItem(str(PERI)))
                form.tabla_2.setItem(i, 6, QTableWidgetItem(str(ID)))
                form.tabla_2.setItem(i, 7, QTableWidgetItem(str(CIND)))
                form.tabla_2.cellWidget(i, 0).cb.setCurrentIndex(0)
                colorea(form.tabla_2, i, len(form.tabla_2.horizontalHeader()), form.tabla_2.cellWidget(i, 0).idx())
    sql = "SELECT temp.REFERENCIA_CATASTRAL AS RC, NIF_SP_VALO AS NIF, temp.ID_VALO, temp.VALORES, NOMBRE_SP_VALO AS SP, temp.IMPORTE AS IMP, " \
          "SWITCH(temp.PER = temp.MINPER, temp.MINPER," \
          "temp.PER <> temp.MINPER, temp.MINPER & '-' & temp.PER) AS PERI " \
          "FROM(" \
          "SELECT NOMBRE_SP_VALO, NIF_SP_VALO, REFERENCIA_CATASTRAL, " \
          "CINT(PERIODO) AS PER FROM [LISTADO 2] ) AS ini " \
          "INNER JOIN (" \
          "SELECT MAX(CINT(PERIODO))AS PER, SUM(IMPORTE_VALO) AS IMPORTE, " \
          "MIN(CINT(PERIODO)) AS MINPER, REFERENCIA_CATASTRAL, FIRST(ID_VALOR) AS ID_VALO, COUNT(ID_VALOR) AS VALORES " \
          "FROM [LISTADO 2] GROUP BY REFERENCIA_CATASTRAL) as temp " \
          "ON (ini.REFERENCIA_CATASTRAL = temp.REFERENCIA_CATASTRAL AND ini.PER= temp.PER);"

    crsr.execute(sql)
    rows = crsr.fetchall()
    for row in rows:
        RC = row.RC
        SP = row.SP
        NIF = row.NIF
        PERI = row.PERI
        IMP = row.IMP
        IMP = float(IMP) / 100
        ID = row.ID_VALO
        if row.VALORES > 1:
            CIND = int(row.VALORES) - 1
        else:
            CIND = ""

        for i in range(form.tabla_2.rowCount()):
            QApplication.processEvents()
            refcat = form.tabla_2.item(i, 5).text()
            if RC == refcat:
                form.tabla_2.setItem(i, 17, QTableWidgetItem(str(SP)))
                form.tabla_2.setItem(i, 18, QTableWidgetItem(str(NIF)))
                form.tabla_2.setItem(i, 12, QTableWidgetItem(str(IMP)))
                form.tabla_2.setItem(i, 8, QTableWidgetItem(str(PERI)))
                form.tabla_2.setItem(i, 6, QTableWidgetItem(str(ID)))
                form.tabla_2.setItem(i, 7, QTableWidgetItem(str(CIND)))
                form.tabla_2.cellWidget(i, 0).cb.setCurrentIndex(0)
                colorea(form.tabla_2, i, len(form.tabla_2.horizontalHeader()), form.tabla_2.cellWidget(i, 0).idx())
    del crsr
    cnxn.close()
    p.close()


def lc():
    print "a"
    anchos = []
    wb = openpyxl.load_workbook(unicode("D:/ahlopez/Escritorio/plantilla.xlsx"))
    print wb
    ws = wb.get_sheet_by_name("ENE-FEB-MAR 17(hasta 10-04) LIQ")
    contador = 0
    for col in ws.columns:
        contador = contador + 1
        arf = (col[0].column, ws.column_dimensions[col[0].column].width)
        anchos.append(arf)
    print anchos


def jj():
    print "a"
    anchos = []
    wb = openpyxl.load_workbook(unicode("D:/ahlopez/Escritorio/fact - copia.xlsx"))
    print wb
    ws = wb.get_sheet_by_name("kilo")
    cell = ws.cell(row=147, column=30)
    print cell.number_format
