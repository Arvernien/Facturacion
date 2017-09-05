# -*- coding: utf-8 -*-

from PyQt4.QtGui import *
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
# ---------------------------------- PARÁMETROS -------------------------------------------------------------
#       Modificar estos valores de ser necesario:
iva = 1.21
preciofinca = 17.43
preciouu16 = 4.61
preciouumas16 = 1.59
# ---------------------------------- PARÁMETROS -------------------------------------------------------------

def confTabla(form):
    form.tabla_2.clear()
    form.tabla_2.setRowCount(0)
    cabeceras = [u'Exp.Carga Catastro', u'Fecha de carga', u'Exp.Alteración', u'Referencia Catastral',
                 u'OBJETO TRIBUTARIO',
                 u'FINCA', u'U.U (Hasta 16 unid.)', u'RESTANTES UU en exceso', u'IMPORTE FINCA',
                 u'IMPORTE U.U (hasta 16 unid.)', u'IMPORTE U.U (resto)', u'FACTURACIÓN', u'OBSERVACIONES']

    form.tabla_2.setColumnCount(len(cabeceras))
    for i in range(form.tabla.rowCount()):
        if form.tabla.cellWidget(i, 0).isChecked():
            row = form.tabla_2.rowCount()
            form.tabla_2.setRowCount(form.tabla_2.rowCount() + 1)
            form.tabla_2.setItem(row, 1, QTableWidgetItem(form.tabla.item(i, 6).text()[-10:]))
            form.tabla_2.setItem(row, 2, QTableWidgetItem(form.tabla.item(i, 9).text()))
            form.tabla_2.setItem(row, 3, QTableWidgetItem(form.tabla.item(i, 12).text()))
            form.tabla_2.setItem(row, 4, QTableWidgetItem(form.tabla.item(i, 15).text()))
            form.tabla_2.setItem(row, 5, QTableWidgetItem(str(1)))
            uu = descomponeUU(int(form.tabla.item(i, 18).text()))
            form.tabla_2.setItem(row, 6, QTableWidgetItem(str(uu[0])))
            form.tabla_2.setItem(row, 7, QTableWidgetItem(str(uu[1])))
            form.tabla_2.setItem(row, 12, QTableWidgetItem(StringCargos(form.tabla.item(i, 17).text())))

    form.tabla_2.setHorizontalHeaderLabels(cabeceras)
    form.tabla_2.horizontalHeader().setResizeMode(QHeaderView.ResizeToContents)


def activarFunciones(form):
    form.desactivarFunciones()
    form.actionImportar_Fichero_de_carga.setEnabled(True)
    form.but_Calcular.setEnabled(True)


def calcular(form):
    for i in range(form.tabla_2.rowCount()):
        fincas = int(form.tabla_2.item(i, 5).text())
        uu16 = int(form.tabla_2.item(i, 6).text())
        uumas16 = int(form.tabla_2.item(i, 7).text())
        factfincas = fincas * preciofinca
        factuu16 = uu16 * preciouu16
        factuumas16 = uumas16 * preciouumas16
        facttotal = factfincas + factuu16 + factuumas16
        form.tabla_2.setItem(i, 8, QTableWidgetItem(str(factfincas)))
        form.tabla_2.setItem(i, 9, QTableWidgetItem(str(factuu16)))
        form.tabla_2.setItem(i, 10, QTableWidgetItem(str(factuumas16)))
        form.tabla_2.setItem(i, 11, QTableWidgetItem(str(facttotal)))


def StringCargos(uu):
    cargos = []
    literalcargos = []
    literalcargos.append(uu.split("|")[0])
    literalcargos.append(uu.split("|")[1])
    for literal in literalcargos:
        previocargos = literal.split(";")
        for a in previocargos:
            if a[:1] == "[":
                intcargos = a[1:-1].split("-")
                for b in range(int(intcargos[0]), int(intcargos[1]) + 1, 1):
                    cargos.append(b)
            else:
                if a != "":
                    cargos.append(a)
    if len(cargos) == 1:
        devolucion = "CARGO "
    else:
        devolucion = "CARGOS "
    for cargo in cargos:
        if cargo != "":
            devolucion = devolucion + str(cargo) + ", "
    return devolucion[:-2]


def descomponeUU(uu):
    a = ['', '']
    if uu >= 16:
        a[0] = 16
        a[1] = uu - 16
    elif uu < 16:
        a[0] = uu
        a[1] = 0
    return a


def importaCarga(form):
    ruta = QFileDialog.getOpenFileName(form, "Seleccionar fichero de carga", "", "Ficheros de carga (*.sc8)")
    if ruta != "":
        file = open(ruta, "r")
        for line in file:
            if line[98:108] == "Expediente":
                expcarga = line[112:124]
            if line[9:10] == ".":
                if line[8:9] != " ":
                    expger = line[1:15]
                    expger = expger.replace(" ", "0")
                refcat = line[47:54] + line[55:62]
                for i in range(form.tabla_2.rowCount()):
                    if form.tabla_2.item(i, 2).text() == expger and form.tabla_2.item(i, 3).text() == refcat:
                        form.tabla_2.setItem(i, 0, QTableWidgetItem(expcarga))


def exportar(form):
    ruta = QFileDialog.getOpenFileName(form, u"Seleccionar fichero de facturación", "", "Ficheros Excel (*.xls *.xlsx)")
    if ruta != "":
        wb = openpyxl.load_workbook(unicode(ruta))
        nhoja, ok = QInputDialog.getText(form, "Nombre de la nueva hoja", "Introduzca nombre:")
        hj = wb.create_sheet(str(nhoja))
        bg = PatternFill(fill_type='solid', start_color='9BBB59', end_color='9BBB59')
        ag = Alignment(horizontal='center', vertical='center')
        borde = Border(left=Side(border_style='thin', color='00000000'),
                       right=Side(border_style='thin', color='00000000'),
                       top=Side(border_style='double', color='00000000'),
                       bottom=Side(border_style='double', color='00000000'))
        fuente = Font(name='Calibri',
                      size=10,
                      bold=True,
                      color='FFFFFFFF')

        for c in range(form.tabla_2.columnCount()):
            d = hj.cell(row=1, column=c + 1, value=str(form.tabla_2.horizontalHeaderItem(c).text()))
            d.fill = bg
            d.alignment = ag
            d.border = borde
            d.font = fuente
        anchos = [('A', 15.42578125), ('B', 13.7109375), ('C', 14.42578125), ('D', 16.7109375), ('E', 40.0), ('F', 9.28515625),
                  ('G', 16.42578125), ('H', 20.5703125), ('I', 13.140625), ('J', 23.85546875), ('K', 17.0), ('L', 12.140625),
                  ('M', 16.42578125), ('N', 11.42578125)]

        for columna, ancho in anchos:
            hj.column_dimensions[columna].width = ancho
        borde = Border(left=Side(border_style='thin', color='00000000'),
                       right=Side(border_style='thin', color='00000000'),
                       top=Side(border_style='thin', color='00000000'),
                       bottom=Side(border_style='thin', color='00000000'))
        fuente = Font(name='Calibri',
                      size=10)
        for fila in range(form.tabla_2.rowCount()):
            for columna in range(form.tabla_2.columnCount()):
                if form.tabla_2.item(fila, columna) is not None:
                    d = hj.cell(row=fila + 2, column=columna + 1, value=str(form.tabla_2.item(fila, columna).text()))
                else:
                    d = hj.cell(row=fila + 2, column=columna + 1, value="")
                d.border = borde
                d.font = fuente
        for col in hj.columns:
            print col[0].column
            if col[0].column == 'M':
                max_length = 0
                column = col[0].column  # Get the column name
                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length*0.75)
                hj.column_dimensions[column].width = adjusted_width
        for col in hj.columns:
            if col[0].column in ['F', 'G', 'H', 'I', 'J', 'K', 'L']:
                for cell in col:
                    try:
                        a = float(cell.value)
                        cell.value = a
                    except:
                        pass
        totalfilas = form.tabla_2.rowCount()
        borde = Border(left=Side(border_style='medium', color='00000000'),
                       right=Side(border_style='medium', color='00000000'),
                       top=Side(border_style='medium', color='00000000'),
                       bottom=Side(border_style='medium', color='00000000'))
        fuente = Font(name='Calibri',
                      size=10,
                      bold=True)
        bg = PatternFill(fill_type='solid', start_color='FABF8F', end_color='FABF8F')

        a = hj.cell(row=totalfilas+2, column=6, value="=SUM(F2:F"+str(totalfilas+1)+")")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 2, column=7, value="=SUM(G2:G" + str(totalfilas + 1) + ")")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 2, column=8, value="=SUM(H2:H" + str(totalfilas + 1) + ")")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 3, column=11, value="TOTAL")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 3, column=12, value="=SUM(L2:L" + str(totalfilas + 1) + ")")
        a.fill = bg
        #a.font = fuente
        a.border = borde
        a.number_format = "0.00"
        a = hj.cell(row=totalfilas + 4, column=11, value="TOTAL CON IVA")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 4, column=12, value="=SUM(L2:L" + str(totalfilas + 1) + ")*"+ str(iva))
        a.fill = bg
        #a.font = fuente
        a.border = borde
        a.number_format = "0.00"
        a = hj.cell(row=totalfilas + 5, column=11, value="IVA")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 5, column=12, value="=L" + str(totalfilas + 4) + "-L" + str(totalfilas + 3))
        a.fill = bg
        #a.font = fuente
        a.border = borde
        a.number_format = "0.00"
        a = hj.cell(row=totalfilas + 6, column=11, value="UU")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 6, column=12, value="=G" + str(totalfilas + 2) + "+H" + str(totalfilas + 2))
        a.fill = bg
        #a.font = fuente
        a.border = borde

        try:
            wb.save(unicode(ruta))
        except IOError as e:
            if e.errno == 13:
                QMessageBox.critical(form, "Error", u"El fichero de destino está en uso")


def listacabeceras():
    print "a"
    anchos = []
    wb = openpyxl.load_workbook(str("D:/ahlopez/Escritorio/murcia fact/FACTURACION catastral Murcia 2017 NUEVA.xlsx"))
    print wb
    ws = wb.get_sheet_by_name("CARGA 07-03-17")
    contador = 0
    for col in ws.columns:
        contador = contador + 1
        arf = (contador, ws.column_dimensions[col[0].column].width)
        anchos.append(arf)
    print anchos

def configura(form):
    pass