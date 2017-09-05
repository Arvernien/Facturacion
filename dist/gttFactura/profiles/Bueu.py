# -*- coding: utf-8 -*-

from PyQt4.QtGui import *
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from datetime import datetime
# ---------------------------------- PARÁMETROS -------------------------------------------------------------
#       Modificar estos valores de ser necesario:
precio = 40
iva = 1.21
# ---------------------------------- PARÁMETROS -------------------------------------------------------------

def confTabla(form):
    form.tabla_2.clear()
    form.tabla_2.setRowCount(0)
    cabeceras = [u'REFERENCIA CATASTRAL', u'CARGO', u'NºEXPEDIENTE', u'EXPTE GERENCIA', u'FECHA ALTERACIÓN', u'FECHA ENTREGA', u'OBJ TRIBUTARIO NUEVO', u'SUJETO PASIVO', u'NIF', u'IMPORTE']

    form.tabla_2.setColumnCount(len(cabeceras))
    for i in range(form.tabla.rowCount()):
        if form.tabla.cellWidget(i, 0).isChecked():

            refcat = form.tabla.item(i, 12).text()
            expger = form.tabla.item(i, 9).text()
            expgtt = form.tabla.item(i, 7).text()
            falteracion = datetime.strptime(str(form.tabla.item(i, 14).text()), '%Y-%m-%d %H:%M:%S').__format__('%d/%m/%Y')
            if form.tabla.item(i, 3).text() != "None":
                fentrega = datetime.strptime(str(form.tabla.item(i, 3).text()), '%Y-%m-%d %H:%M:%S').__format__('%d/%m/%Y')
            else:
                fentrega = ""
            objtributario = form.tabla.item(i, 15).text()

            lcargos = DevuelveCargos(form.tabla.item(i, 17).text())
            for i in range(0, len(lcargos), 1):
                print refcat, componeCargo(lcargos[i])
                row = form.tabla_2.rowCount()
                form.tabla_2.setRowCount(form.tabla_2.rowCount() + 1)
                form.tabla_2.setItem(row, 0, QTableWidgetItem(refcat))
                form.tabla_2.setItem(row, 1, QTableWidgetItem(componeCargo(lcargos[i])))
                form.tabla_2.setItem(row, 2, QTableWidgetItem(expgtt))
                form.tabla_2.setItem(row, 3, QTableWidgetItem(expger))
                form.tabla_2.setItem(row, 4, QTableWidgetItem(str(falteracion)))
                form.tabla_2.setItem(row, 5, QTableWidgetItem(str(fentrega)))
                form.tabla_2.setItem(row, 6, QTableWidgetItem(objtributario))
                form.tabla_2.setItem(row, 9, QTableWidgetItem(str(precio)))
            # form.tabla_2.setItem(row, 5, QTableWidgetItem(str(1)))
            # uu = descomponeUU(int(form.tabla.item(i, 18).text()))
            # form.tabla_2.setItem(row, 6, QTableWidgetItem(str(uu[0])))
            # form.tabla_2.setItem(row, 7, QTableWidgetItem(str(uu[1])))
            # form.tabla_2.setItem(row, 12, QTableWidgetItem(StringCargos(form.tabla.item(i, 17).text())))

    form.tabla_2.setHorizontalHeaderLabels(cabeceras)
    form.tabla_2.horizontalHeader().setResizeMode(QHeaderView.ResizeToContents)

def componeCargo(cargo):
    cargo = str(cargo)
    while len(cargo) < 4:
        cargo = "0"+cargo
    return cargo
def activarFunciones(form):
    form.desactivarFunciones()
    form.actionImportar_datos_desde_FIN_de_Retorno.setEnabled(True)
    form.actionImportar_datos_desde_FIN_de_Retorno.setText("Importar FIN de Salida")

def configura(form):
    pass


def calcular(form):
    for i in range(form.tabla_2.rowCount()):
        fincas = int(form.tabla_2.item(i, 5).text())
        uu16 = int(form.tabla_2.item(i, 6).text())
        uumas16 = int(form.tabla_2.item(i, 7).text())
        factfincas = fincas * preciofinca
        factuu16 = uu16 * preciouu16
        factuumas16 = uumas16 * preciouumas16
        facttotal = factfincas + factuu16 + factuumas16
        form.tabla_2.setItem(i, 8, QTableWidgetItem(str(factfincas)))
        form.tabla_2.setItem(i, 9, QTableWidgetItem(str(factuu16)))
        form.tabla_2.setItem(i, 10, QTableWidgetItem(str(factuumas16)))
        form.tabla_2.setItem(i, 11, QTableWidgetItem(str(facttotal)))

def importaFin(form):
    ruta = QFileDialog.getOpenFileName(form, "Seleccionar fichero FIN", "", "FIN RETORNO (*.*)")
    if ruta != "":
        file = open(unicode(ruta), "r")
        p = form.loading("Leyendo FIN", form.tabWidget)
        p.show()
        for line in file:
            QApplication.processEvents()
            if line[0:2] == "01":
                print line[53:57]
                tipofin = line[53:57]
                print tipofin != "CFIS"
                if tipofin != "CFIS":
                    QMessageBox.critical(form, "Error", u"El fichero no es un fichero FIN de salida válido")
                    return
            if line[0:2] == "15":
                obj = line[200:205].strip() + " " + line[205:230].strip() + " " + str(int(line[230:234])) + line[234:235]
                refcat = line[30:44]
                cargo = line[44:48]
                escalera = line[249:251].strip()
                planta = line[251:254].strip()
                puerta = line[254:257].strip()
                if escalera != "":
                    escalera = " ES:" + escalera
                if planta != "":
                    planta = " PL:" + planta
                if puerta != "":
                    puerta = " PT:" + puerta
                obj = obj + escalera + planta + puerta
                for i in range(form.tabla_2.rowCount()):
                    if form.tabla_2.item(i, 0).text() == refcat and form.tabla_2.item(i, 1).text() == cargo:
                        form.tabla_2.setItem(i, 6, QTableWidgetItem(obj))

            if line[0:2] == "46":
                if line[57:60] == "001":
                    refcat = line[30:44]
                    cargo = line[44:48]
                    titular = line[69:129]
                    titular = titular.strip()
                    nif = line[60:69]
                    for i in range(form.tabla_2.rowCount()):
                        if form.tabla_2.item(i, 0).text() == refcat and form.tabla_2.item(i, 1).text() == cargo:
                            form.tabla_2.setItem(i, 6, QTableWidgetItem(obj))
                            form.tabla_2.setItem(i, 7, QTableWidgetItem(titular))
                            form.tabla_2.setItem(i, 8, QTableWidgetItem(nif))
        p.close()

def StringCargos(uu):
    cargos = []
    literalcargos = []
    literalcargos.append(uu.split("|")[0])
    literalcargos.append(uu.split("|")[1])
    for literal in literalcargos:
        previocargos = literal.split(";")
        for a in previocargos:
            if a[:1] == "[":
                intcargos = a[1:-1].split("-")
                for b in range(int(intcargos[0]), int(intcargos[1]) + 1, 1):
                    cargos.append(b)
            else:
                if a != "":
                    cargos.append(a)
    if len(cargos) == 1:
        devolucion = "CARGO "
    else:
        devolucion = "CARGOS "
    for cargo in cargos:
        if cargo != "":
            devolucion = devolucion + str(cargo) + ", "
    return devolucion[:-2]

def DevuelveCargos(uu):
    cargos = []
    literalcargos = []
    literalcargos.append(uu.split("|")[0])
    literalcargos.append(uu.split("|")[1])
    for literal in literalcargos:
        previocargos = literal.split(";")
        for a in previocargos:
            if a[:1] == "[":
                intcargos = a[1:-1].split("-")
                for b in range(int(intcargos[0]), int(intcargos[1]) + 1, 1):
                    cargos.append(b)
            else:
                if a != "":
                    cargos.append(a)
    return cargos


def descomponeUU(uu):
    a = ['', '']
    if uu >= 16:
        a[0] = 16
        a[1] = uu - 16
    elif uu < 16:
        a[0] = uu
        a[1] = 0
    return a

def exportar(form):
    ruta = QFileDialog.getOpenFileName(form, u"Seleccionar fichero de facturación", "", "Ficheros Excel (*.xls *.xlsx)")
    if ruta != "":
        wb = openpyxl.load_workbook(unicode(ruta))
        nhoja, ok = QInputDialog.getText(form, "Nombre de la nueva hoja", "Introduzca nombre:")
        hj = wb.create_sheet(str(nhoja))
        bg = PatternFill(fill_type='solid', start_color='FABF8F', end_color='FABF8F')
        ag = Alignment(horizontal='center', vertical='center')
        borde = Border(left=Side(border_style='thin', color='000000'),
                       right=Side(border_style='thin', color='000000'),
                       top=Side(border_style='thin', color='000000'),
                       bottom=Side(border_style='thin', color='000000'))
        fuente = Font(name='Calibri',
                      size=10,
                      bold=True,
                      color='000000')

        for c in range(form.tabla_2.columnCount()):
            d = hj.cell(row=1, column=c + 1, value=str(form.tabla_2.horizontalHeaderItem(c).text()))
            d.fill = bg
            d.alignment = ag
            d.border = borde
            d.font = fuente
        anchos = [('A', 22.42578125), ('B', 5.85546875), ('C', 15.0), ('D', 14.140625), ('E', 13.5703125),
                  ('F', 16.42578125), ('G', 66.5703125), ('H', 45.28515625), ('I', 12.5703125), ('J', 8.85546875)]


        for columna, ancho in anchos:
            hj.column_dimensions[columna].width = ancho
        borde = Border(left=Side(border_style='thin', color='00000000'),
                       right=Side(border_style='thin', color='00000000'),
                       top=Side(border_style='thin', color='00000000'),
                       bottom=Side(border_style='thin', color='00000000'))
        fuente = Font(name='Calibri',
                      size=10)
        for fila in range(form.tabla_2.rowCount()):
            for columna in range(form.tabla_2.columnCount()):
                if form.tabla_2.item(fila, columna) is not None:
                    d = hj.cell(row=fila + 2, column=columna + 1, value=str(form.tabla_2.item(fila, columna).text()))
                else:
                    d = hj.cell(row=fila + 2, column=columna + 1, value="")
                d.border = borde
                d.font = fuente
        # for col in hj.columns:
        #     print col[0].column
        #     if col[0].column in ['H']:
        #         max_length = 0
        #         column = col[0].column  # Get the column name
        #         for cell in col:
        #             try:  # Necessary to avoid error on empty cells
        #                 if len(str(cell.value)) > max_length:
        #                     max_length = len(cell.value)
        #             except:
        #                 pass
        #         adjusted_width = (max_length*1)
        #         hj.column_dimensions[column].width = adjusted_width
        for col in hj.columns:
            if col[0].column in ['J']:
                for cell in col:
                    try:
                        a = float(cell.value)
                        cell.value = a
                        cell.number_format = "0.00"
                    except:
                        pass
        totalfilas = form.tabla_2.rowCount()
        borde = Border(left=Side(border_style='medium', color='00000000'),
                       right=Side(border_style='medium', color='00000000'),
                       top=Side(border_style='medium', color='00000000'),
                       bottom=Side(border_style='medium', color='00000000'))
        fuente = Font(name='Calibri',
                      size=10,
                      bold=True)
        bg = PatternFill(fill_type='solid', start_color='FABF8F', end_color='FABF8F')

        a = hj.cell(row=totalfilas + 3, column=9, value="TOTAL")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 3, column=10, value="=SUM(J2:J" + str(totalfilas + 1) + ")")
        a.fill = bg
        #a.font = fuente
        a.border = borde
        a.number_format = "0.00"
        a = hj.cell(row=totalfilas + 4, column=9, value="TOTAL CON IVA")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 4, column=10, value="=SUM(J2:J" + str(totalfilas + 1) + ")*" + str(iva))
        a.fill = bg
        #a.font = fuente
        a.border = borde
        a.number_format = "0.00"
        a = hj.cell(row=totalfilas + 5, column=9, value="IVA")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 5, column=10, value="=J" + str(totalfilas + 4) + "-J" + str(totalfilas + 3))
        a.fill = bg
        #a.font = fuente
        a.border = borde
        a.number_format = "0.00"
        a = hj.cell(row=totalfilas + 6, column=9, value="UU")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 6, column=10, value="=" + str(totalfilas))
        a.fill = bg
        #a.font = fuente
        a.border = borde
        a.number_format = "0"

        try:
            wb.save(unicode(ruta))
            QMessageBox.information(form, "Guardado", u"Guardado correctamente")
        except IOError as e:
            if e.errno == 13:
                QMessageBox.critical(form, "Error", u"El fichero de destino está en uso")


def listacabeceras():
    print "a"
    anchos = []
    wb = openpyxl.load_workbook(str("D:/ahlopez/Escritorio/FACTURACION BUEU 2017.xlsx"))
    print wb
    ws = wb.get_sheet_by_name("CARGA 13-02-17")
    contador = 0
    for col in ws.columns:
        contador = contador + 1
        arf = (chr(64+contador), ws.column_dimensions[col[0].column].width)
        anchos.append(arf)
    print anchos