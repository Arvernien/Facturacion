# -*- coding: utf-8 -*-

from PyQt4.QtGui import *
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
# ---------------------------------- PARÁMETROS -------------------------------------------------------------
#       Modificar estos valores de ser necesario:
iva = 1.21              # Porcentaje IVA.
preciouu1 = 31.03       # Precio de 1 UU.
preciouu2a10 = 25.86    # Precio de 2 a 10 UU.
preciouu11a50 = 20.69   # Precio de 11 a 50 UU.
preciouu51a100 = 15.52  # Precio de 51 a 100 UU.
preciouu101 = 8.62      # Precio de más de 101 UU.
preciorecurso = 25.86   # Precio de UU en recurso.

# ---------------------------------- PARÁMETROS -------------------------------------------------------------

def confTabla(form):
    form.tabla_2.clear()
    form.tabla_2.setRowCount(0)
    cabeceras = [u'REF. CATASTRAL', u'UNIDADES', u'TIPO DE ACTUACIÓN', u'TIPO EXPTE',
                 u'EXP. AYTO',    u'EXP. GERENCIA', u'1 UU', u'2 A 10 UU', u'11 A 50 UU',
                 u'51 A 100 UU', u'101+ UU', u'UU RECURSO', u'FACTURACIÓN', u'OBSERVACIONES']


    form.tabla_2.setColumnCount(len(cabeceras))
    for i in range(form.tabla.rowCount()):
        if form.tabla.cellWidget(i, 0).isChecked():
            refcat = form.tabla.item(i, 12).text()
            expger = form.tabla.item(i, 9).text()
            expgtt = form.tabla.item(i, 7).text()
            uu = form.tabla.item(i, 18).text()
            actuacion = form.tabla.item(i, 19).text()
            tipoexp = form.tabla.item(i, 20).text()
            row = form.tabla_2.rowCount()
            form.tabla_2.setRowCount(form.tabla_2.rowCount() + 1)
            form.tabla_2.setItem(row, 0, QTableWidgetItem(refcat))
            form.tabla_2.setItem(row, 1, QTableWidgetItem(uu))
            form.tabla_2.setItem(row, 2, QTableWidgetItem(actuacion))
            form.tabla_2.setItem(row, 3, QTableWidgetItem(tipoexp))
            form.tabla_2.setItem(row, 4, QTableWidgetItem(expgtt))
            form.tabla_2.setItem(row, 5, QTableWidgetItem(expger))
            form.tabla_2.setItem(row, 13, QTableWidgetItem(StringCargos(form.tabla.item(i, 17).text())))

    form.tabla_2.setHorizontalHeaderLabels(cabeceras)
    form.tabla_2.horizontalHeader().setResizeMode(QHeaderView.ResizeToContents)

def configura(form):
    pass


def activarFunciones(form):
    form.desactivarFunciones()
    form.actionImportar_Fichero_de_carga.setEnabled(True)
    form.but_Calcular.setEnabled(True)


def calcular(form):
    for i in range(form.tabla_2.rowCount()):
        limpiafila(i, form)
        uu = int(form.tabla_2.item(i, 1).text())
        tipo = form.tabla_2.item(i, 2).text()
        if tipo == "RECURSO":
            columna = 11
            fact = uu * preciorecurso
        else:
            print uu, uu < 2
            if uu == 1:
                columna = 6
                fact = uu * preciouu1
            elif 2 <= uu < 11:
                columna = 7
                fact = uu * preciouu2a10
            elif 11 <= uu < 51:
                columna = 8
                fact = uu * preciouu11a50
            elif 51 <= uu < 101:
                columna = 9
                fact = uu * preciouu51a100
            elif uu > 100:
                columna = 10
                fact = uu * preciouu101
        form.tabla_2.setItem(i, columna, QTableWidgetItem(str(uu)))
        form.tabla_2.setItem(i, 12, QTableWidgetItem(str(fact)))


        # factfincas = fincas * preciofinca
        # factuu16 = uu16 * preciouu16
        # factuumas16 = uumas16 * preciouumas16
        # facttotal = factfincas + factuu16 + factuumas16
        # form.tabla_2.setItem(i, 8, QTableWidgetItem(str(factfincas)))
        # form.tabla_2.setItem(i, 9, QTableWidgetItem(str(factuu16)))
        # form.tabla_2.setItem(i, 10, QTableWidgetItem(str(factuumas16)))
        # form.tabla_2.setItem(i, 11, QTableWidgetItem(str(facttotal)))

def limpiafila(i, form):
    for a in range(6, 11, 1):
        form.tabla_2.setItem(i, a, QTableWidgetItem(""))

def StringCargos(uu):
    lista = []
    modificados = []
    altas = []
    bajas = []

    #uu.split("|")[1]
    literal = uu.split("|")[0]
    previocargos = literal.split(";")
    for a in previocargos:
        if a[:1] == "[":
            intcargos = a[1:-1].split("-")
            for b in range(int(intcargos[0]), int(intcargos[1]) + 1, 1):
                modificados.append(b)
        else:
            if a != "":
                modificados.append(a)
    strmod = "Mod: "
    for cargo in modificados:
        if cargo != "":
            strmod = strmod + str(cargo) + ", "
    if len(modificados) > 0:
        lista.append(strmod[:-2] + ".")



    literal = uu.split("|")[1]
    previocargos = literal.split(";")
    for a in previocargos:
        if a[:1] == "[":
            intcargos = a[1:-1].split("-")
            for b in range(int(intcargos[0]), int(intcargos[1]) + 1, 1):
                altas.append(b)
        else:
            if a != "":
                altas.append(a)
    stralta = "Altas: "
    for cargo in altas:
        if cargo != "":
            stralta = stralta + str(cargo) + ", "
    if len(altas) > 0:
        lista.append(stralta[:-2] + ".")

    literal = uu.split("|")[2]
    previocargos = literal.split(";")
    for a in previocargos:
        if a[:1] == "[":
            intcargos = a[1:-1].split("-")
            for b in range(int(intcargos[0]), int(intcargos[1]) + 1, 1):
                bajas.append(b)
        else:
            if a != "":
                bajas.append(a)
    strbaja = "Bajas: "
    for cargo in bajas:
        if cargo != "":
            strbaja = strbaja + str(cargo) + ", "
    if len(bajas) > 0:
        lista.append(strbaja[:-2] + ".")
    return " ".join(lista)


def descomponeUU(uu):
    a = ['', '']
    if uu >= 16:
        a[0] = 16
        a[1] = uu - 16
    elif uu < 16:
        a[0] = uu
        a[1] = 0
    return a


def importaCarga(form):
    ruta = QFileDialog.getOpenFileName(form, "Seleccionar fichero de carga", "", "Ficheros de carga (*.sc8)")
    if ruta != "":
        file = open(ruta, "r")
        for line in file:
            if line[98:108] == "Expediente":
                expcarga = line[112:124]
            if line[9:10] == ".":
                if line[8:9] != " ":
                    expger = line[1:15]
                    expger = expger.replace(" ", "0")
                refcat = line[47:54] + line[55:62]
                for i in range(form.tabla_2.rowCount()):
                    if form.tabla_2.item(i, 2).text() == expger and form.tabla_2.item(i, 3).text() == refcat:
                        form.tabla_2.setItem(i, 0, QTableWidgetItem(expcarga))


def exportar(form):
    ruta = QFileDialog.getOpenFileName(form, u"Seleccionar fichero de facturación", "", "Ficheros Excel (*.xls *.xlsx)")
    if ruta != "":
        wb = openpyxl.load_workbook(unicode(ruta))
        nhoja, ok = QInputDialog.getText(form, "Nombre de la nueva hoja", "Introduzca nombre:")
        hj = wb.create_sheet(str(nhoja))
        bg = PatternFill(fill_type='solid', start_color='A5A5A5', end_color='A5A5A5')
        ag = Alignment(horizontal='center', vertical='center')
        borde = Border(left=Side(border_style='thin', color='00000000'),
                       right=Side(border_style='thin', color='00000000'),
                       top=Side(border_style='double', color='00000000'),
                       bottom=Side(border_style='double', color='00000000'))
        fuente = Font(name='Calibri',
                      size=10,
                      bold=True,
                      color='FFFFFFFF')

        for c in range(form.tabla_2.columnCount()):
            d = hj.cell(row=1, column=c + 1, value=str(form.tabla_2.horizontalHeaderItem(c).text()))
            d.fill = bg
            d.alignment = ag
            d.border = borde
            d.font = fuente
        anchos = [('A', 17.42578125), ('B', 12.42578125), ('C', 25.140625), ('D', 10.85546875), ('E', 14.7109375),
                  ('F', 15.42578125), ('G', 11.42578125), ('H', 11.42578125), ('I', 11.42578125), ('J', 11.42578125),
                  ('K', 12.5703125), ('L', 12.5703125), ('M', 13.85546875), ('N', 18.140625)]


        for columna, ancho in anchos:
            hj.column_dimensions[columna].width = ancho
        borde = Border(left=Side(border_style='thin', color='00000000'),
                       right=Side(border_style='thin', color='00000000'),
                       top=Side(border_style='thin', color='00000000'),
                       bottom=Side(border_style='thin', color='00000000'))
        fuente = Font(name='Calibri',
                      size=10)
        for fila in range(form.tabla_2.rowCount()):
            for columna in range(form.tabla_2.columnCount()):
                if form.tabla_2.item(fila, columna) is not None:
                    d = hj.cell(row=fila + 2, column=columna + 1, value=str(form.tabla_2.item(fila, columna).text()))
                else:
                    d = hj.cell(row=fila + 2, column=columna + 1, value="")
                d.border = borde
                d.font = fuente
        for col in hj.columns:
            if col[0].column == 'N':
                max_length = 0
                column = col[0].column  # Get the column name
                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length*0.85)
                hj.column_dimensions[column].width = adjusted_width
        for col in hj.columns:
            if col[0].column in ['B', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                for cell in col:
                    try:
                        a = float(cell.value)
                        cell.value = a
                    except:
                        pass
        for col in hj.columns:
            if col[0].column in ['B', 'G', 'H', 'I', 'J', 'K', 'L']:
                for cell in col:
                    try:
                        cell.number_format = "0"
                    except:
                        pass
        totalfilas = form.tabla_2.rowCount()
        borde = Border(left=Side(border_style='medium', color='00000000'),
                       right=Side(border_style='medium', color='00000000'),
                       top=Side(border_style='medium', color='00000000'),
                       bottom=Side(border_style='medium', color='00000000'))
        fuente = Font(name='Calibri',
                      size=10,
                      bold=True)
        bg = PatternFill(fill_type='solid', start_color='FABF8F', end_color='FABF8F')


        a = hj.cell(row=totalfilas + 2, column=7, value="=SUM(G2:G" + str(totalfilas + 1) + ")")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 2, column=8, value="=SUM(H2:H" + str(totalfilas + 1) + ")")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 2, column=9, value="=SUM(I2:I" + str(totalfilas + 1) + ")")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 2, column=10, value="=SUM(J2:J" + str(totalfilas + 1) + ")")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 2, column=11, value="=SUM(K2:K" + str(totalfilas + 1) + ")")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 2, column=12, value="=SUM(L2:L" + str(totalfilas + 1) + ")")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 4, column=12, value="TOTAL")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 4, column=13, value="=SUM(M2:M" + str(totalfilas + 1) + ")")
        a.fill = bg
        #a.font = fuente
        a.border = borde
        a.number_format = "0.00"
        a = hj.cell(row=totalfilas + 5, column=12, value="TOTAL CON IVA")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 5, column=13, value="=SUM(M2:M" + str(totalfilas + 1) + ")*"+ str(iva))
        a.fill = bg
        #a.font = fuente
        a.border = borde
        a.number_format = "0.00"
        a = hj.cell(row=totalfilas + 6, column=12, value="IVA")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 6, column=13, value="=M" + str(totalfilas + 4) + "-M" + str(totalfilas + 3))
        a.fill = bg
        #a.font = fuente
        a.border = borde
        a.number_format = "0.00"
        a = hj.cell(row=totalfilas + 7, column=12, value="UU")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 7, column=13, value="=SUM(G" + str(totalfilas + 2) + ":L" + str(totalfilas + 2) + ")")
        a.fill = bg
        #a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 8, column=12, value="PRECIO UU")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 8, column=13, value="=M" + str(totalfilas + 4) + "/M" + str(totalfilas + 7))
        a.fill = bg
        # a.font = fuente
        a.border = borde
        a.number_format = "0.00"
        bg = PatternFill(fill_type='solid', start_color='A9D08E', end_color='A9D08E')
        a = hj.cell(row=totalfilas + 4, column=1, value="UNIDADES URBANAS")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 4, column=2, value="PRECIO")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a.number_format = "0.00"
        borde = Border(left=Side(border_style='thin', color='00000000'),
                       right=Side(border_style='thin', color='00000000'),
                       top=Side(border_style='thin', color='00000000'),
                       bottom=Side(border_style='thin', color='00000000'))
        fuente = Font(name='Calibri',
                      size=10,
                      bold=False)
        a = hj.cell(row=totalfilas + 5, column=1, value="1 UNIDAD")
        a.font = fuente

        a.border = borde
        a = hj.cell(row=totalfilas + 5, column=2, value=str(preciouu1))
        a.font = fuente

        a.border = borde
        a.number_format = "0.00"
        a = hj.cell(row=totalfilas + 6, column=1, value="DE 2 A 10")
        a.font = fuente

        a.border = borde
        a = hj.cell(row=totalfilas + 6, column=2, value=str(preciouu2a10))
        a.font = fuente

        a.border = borde
        a.number_format = "0.00"
        a = hj.cell(row=totalfilas + 7, column=1, value="DE 11 A 50")
        a.font = fuente

        a.border = borde
        a = hj.cell(row=totalfilas + 7, column=2, value=str(preciouu11a50))
        a.font = fuente

        a.border = borde
        a.number_format = "0.00"
        a = hj.cell(row=totalfilas + 8, column=1, value="DE 51 A 100")
        a.font = fuente

        a.border = borde
        a = hj.cell(row=totalfilas + 8, column=2, value=str(preciouu51a100))
        a.font = fuente

        a.border = borde
        a.number_format = "0.00"
        a = hj.cell(row=totalfilas + 9, column=1, value=U"101 Y MÁS")
        a.font = fuente

        a.border = borde
        a = hj.cell(row=totalfilas + 9, column=2, value=str(preciouu101))
        a.font = fuente

        a.border = borde
        a.number_format = "0.00"
        a = hj.cell(row=totalfilas + 10, column=1, value="RECURSOS")
        a.font = fuente

        a.border = borde
        a = hj.cell(row=totalfilas + 10, column=2, value=str(preciorecurso))
        a.font = fuente

        a.border = borde
        a.number_format = "0.00"

        try:
            wb.save(unicode(ruta))
        except IOError as e:
            if e.errno == 13:
                QMessageBox.critical(form, "Error", u"El fichero de destino está en uso")


def listacabeceras():
    print "a"
    anchos = []
    wb = openpyxl.load_workbook(u"D:/ahlopez/Escritorio/FACTURACIÓN 2016.xlsx")
    print wb
    ws = wb.get_sheet_by_name("Entrega julio")
    contador = 0
    for col in ws.columns:
        contador = contador + 1
        arf = (contador, ws.column_dimensions[col[0].column].width)
        anchos.append(arf)
    print anchos