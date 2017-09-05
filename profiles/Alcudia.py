# -*- coding: utf-8 -*-

from PyQt4.QtGui import *
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from datetime import datetime, date
import time
import pyodbc
import os

# ---------------------------------- PARÁMETROS -------------------------------------------------------------
#       Modificar estos valores de ser necesario:
precio = 40
iva = 0.21
minoar = 51.00
mingtt = 44.00
# ---------------------------------- PARÁMETROS -------------------------------------------------------------

def confTabla(form):
    form.tabla_2.clear()
    form.tabla_2.setRowCount(0)
    y = date.today().year
    cabeceras = [u'Ref. Catastral', u'Exp. Gerencia', u'Nº fijo', u'DIRECCIÓN OT', u'NIF', u'SUJETO PASIVO', u'Codigo Postal',
                 u'Nom. Provincia',
                 u'Nom.Municipio', u'Clave Uso', u'IDVALOR', u'Periodo', u'Liquidac.1', u'Liquidac.2',
                 u'liquidac. A deducir',
                 u'Total Liquidac.', u'4,5% liq', u'ValorCat ' + str(y), u'BaseLiq ' + str(y),
                 u'BaseLiquidable ' + str(y + 1),
                 u'BaseLiq a deducir', u'Cuota ' + str(y + 1), u'4,5% inc padrón ' + str(y + 1), u'agrup/segr/otros',
                 u'total liq o total incorp', u'FACTURACIÓN', u'OBSERVACIONES']

    form.tabla_2.setColumnCount(len(cabeceras))
    for i in range(form.tabla.rowCount()):
        if form.tabla.cellWidget(i, 0).isChecked():

            refcat = form.tabla.item(i, 12).text()
            expger = form.tabla.item(i, 9).text()
            expgtt = form.tabla.item(i, 7).text()
            municipio = form.tabla.item(i, 5).text()
            id_subo = form.tabla.item(i, 4).text()
            falteracion = datetime.strptime(str(form.tabla.item(i, 14).text()), '%Y-%m-%d %H:%M:%S').__format__(
                '%d/%m/%Y')
            if form.tabla.item(i, 3).text() != "None":
                fentrega = datetime.strptime(str(form.tabla.item(i, 3).text()), '%Y-%m-%d %H:%M:%S').__format__(
                    '%d/%m/%Y')
            else:
                fentrega = ""
            objtributario = form.tabla.item(i, 15).text()

            lcargos = DevuelveCargos(form.tabla.item(i, 17).text())
            for i in range(0, len(lcargos), 1):
                QApplication.processEvents()
                row = form.tabla_2.rowCount()
                form.tabla_2.setRowCount(form.tabla_2.rowCount() + 1)
                form.tabla_2.setItem(row, 0, QTableWidgetItem(refcat + componeCargo(lcargos[i])))
                form.tabla_2.setItem(row, 1, QTableWidgetItem(expger))
                form.tabla_2.setItem(row, 7, QTableWidgetItem("VALENCIA"))
                form.tabla_2.setItem(row, 8, QTableWidgetItem(municipio))

    form.tabla_2.setHorizontalHeaderLabels(cabeceras)
    form.tabla_2.horizontalHeader().setResizeMode(QHeaderView.ResizeToContents)


def componeCargo(cargo):
    cargo = str(cargo)
    while len(cargo) < 4:
        cargo = "0" + cargo
    return cargo


def activarFunciones(form):
    form.desactivarFunciones()
    form.actionImportar_datos_desde_FIN_de_Retorno.setEnabled(True)
    form.actionImportar_datos_desde_FIN_de_Retorno.setText("Importar DOC")
    form.but_Calcular.setEnabled(False)
    form.actionImportar_Access_Valores.setVisible(False)


def configura(form):
    pass


def calcular(form):
    for i in range(form.tabla_2.rowCount()):
        QApplication.processEvents()
        if form.tabla_2.item(i, 21) is None or (
                        form.tabla_2.item(i, 21) is not None and form.tabla_2.item(i, 21).text() == ""):
            if form.tabla_2.item(i, 20).text() != "MANUAL":
                bliq = float(form.tabla_2.item(i, 19).text())
                tipo = float(form.tabla_2.item(i, 20).text())
                pad = bliq * tipo / 100
                padoar025 = pad * 0.25
                pad2050gtt = pad * 0.205
                form.tabla_2.setItem(i, 0, QTableWidgetItem("3"))
                form.tabla_2.setItem(i, 4, QTableWidgetItem("N"))
                form.tabla_2.setItem(i, 16, QTableWidgetItem(str(time.strftime('%Y'))))
                form.tabla_2.setItem(i, 24, QTableWidgetItem(str(round(pad, 2))))
                form.tabla_2.setItem(i, 25, QTableWidgetItem(str(round(pad, 2))))
                form.tabla_2.setItem(i, 26, QTableWidgetItem(str(round(padoar025, 2))))
                form.tabla_2.setItem(i, 27, QTableWidgetItem(str(round(pad2050gtt, 2))))
                form.tabla_2.setItem(i, 28, QTableWidgetItem(str(round(padoar025, 2))))
                if "AYUNTAMIENTO" not in form.tabla_2.item(i, 18).text():
                    if padoar025 < minoar:
                        totaloar = minoar
                    else:
                        totaloar = padoar025
                else:
                    totaloar = minoar
                form.tabla_2.setItem(i, 29, QTableWidgetItem(str(round(totaloar, 2))))
                form.tabla_2.setItem(i, 30, QTableWidgetItem(str(round(totaloar, 2))))
                form.tabla_2.setItem(i, 31, QTableWidgetItem(str(round(pad2050gtt, 2))))
                if "AYUNTAMIENTO" not in form.tabla_2.item(i, 18).text():
                    if pad2050gtt < mingtt:
                        totalgtt = mingtt
                    else:
                        totalgtt = pad2050gtt
                else:
                    totalgtt = mingtt
                form.tabla_2.setItem(i, 32, QTableWidgetItem(str(round(totalgtt, 2))))
                form.tabla_2.setItem(i, 33, QTableWidgetItem(str(round(totalgtt, 2))))
                form.tabla_2.setItem(i, 22, QTableWidgetItem(""))
                form.tabla_2.setItem(i, 23, QTableWidgetItem(""))
        else:
            bliq = float(form.tabla_2.item(i, 19).text())
            tipo = float(form.tabla_2.item(i, 20).text())
            pad = bliq * tipo / 100
            liq = float(form.tabla_2.item(i, 21).text())
            liqoar = liq * 0.25
            liqgtt = liq * 0.2050
            form.tabla_2.setItem(i, 0, QTableWidgetItem("1"))
            form.tabla_2.setItem(i, 4, QTableWidgetItem("S"))
            form.tabla_2.setItem(i, 22, QTableWidgetItem(str(round(liqoar, 2))))
            form.tabla_2.setItem(i, 23, QTableWidgetItem(str(round(liqgtt, 2))))
            form.tabla_2.setItem(i, 24, QTableWidgetItem(str(round(pad, 2))))
            form.tabla_2.setItem(i, 28, QTableWidgetItem(str(round(liqoar, 2))))
            if "AYUNTAMIENTO" not in form.tabla_2.item(i, 18).text():
                if liqoar < minoar:
                    totaloar = minoar
                else:
                    totaloar = liqoar
            else:
                totaloar = minoar
            form.tabla_2.setItem(i, 29, QTableWidgetItem(str(round(totaloar, 2))))
            form.tabla_2.setItem(i, 30, QTableWidgetItem(str(round(totaloar, 2))))
            form.tabla_2.setItem(i, 31, QTableWidgetItem(str(round(liqgtt, 2))))
            if "AYUNTAMIENTO" not in form.tabla_2.item(i, 18).text():
                if liqgtt < mingtt:
                    totalgtt = mingtt
                else:
                    totalgtt = liqgtt
            else:
                totalgtt = mingtt
            form.tabla_2.setItem(i, 32, QTableWidgetItem(str(round(totalgtt, 2))))
            form.tabla_2.setItem(i, 33, QTableWidgetItem(str(round(totalgtt, 2))))
            form.tabla_2.setItem(i, 25, QTableWidgetItem(""))
            form.tabla_2.setItem(i, 26, QTableWidgetItem(""))
            form.tabla_2.setItem(i, 27, QTableWidgetItem(""))


def importaFin(form):
    ruta = QFileDialog.getOpenFileName(form, "Seleccionar fichero FIN", "", "FIN RETORNO (*.*)")
    if ruta != "":
        file = open(unicode(ruta), "r")
        p = form.loading("Leyendo DOC", form.tabWidget)
        p.show()
        for line in file:
            QApplication.processEvents()
            if line[0:2] == "01":
                print line[53:57]
                tipofin = line[53:57]
            if line[0:2] == "53":
                obj = line[160:165].strip() + " " + line[165:190].strip() + " " + str(int(line[190:194])) + line[
                                                                                                            194:195]
                refcat = line[9:29]
                expediente = line[821:829] + "." + line[830:832] + "/" + line[817:819]
                tipoexp = line[803:807]
                escalera = line[687:689].strip()
                planta = line[689:692].strip()
                puerta = line[692:695].strip()
                nfijo = line[29:37]
                vcat = str(float(line[296:308]) / 100)
                bliq = str(float(line[332:344]) / 100)
                titular = line[462:522].strip()
                nif = line[453:462].strip()
                cpost = line[242:247].strip()
                uso = line[344:345]
                if escalera != "":
                    escalera = " ES:" + escalera
                if planta != "":
                    planta = " PL:" + planta
                if puerta != "":
                    puerta = " PT:" + puerta
                obj = obj + escalera + planta + puerta
                for i in range(form.tabla_2.rowCount()):
                    if form.tabla_2.item(i, 0).text()[:18] == refcat[:18]:
                        form.tabla_2.setItem(i, 0, QTableWidgetItem(refcat))
                        form.tabla_2.setItem(i, 3, QTableWidgetItem(obj))
                        form.tabla_2.setItem(i, 6, QTableWidgetItem(cpost))
                        form.tabla_2.setItem(i, 2, QTableWidgetItem(nfijo))
                        form.tabla_2.setItem(i, 17, QTableWidgetItem(vcat))
                        form.tabla_2.setItem(i, 18, QTableWidgetItem(bliq))
                        form.tabla_2.setItem(i, 19, QTableWidgetItem(bliq))
                        form.tabla_2.setItem(i, 4, QTableWidgetItem(nif))
                        form.tabla_2.setItem(i, 5, QTableWidgetItem(titular))
                        form.tabla_2.setItem(i, 9, QTableWidgetItem(uso))
        p.close()


def StringCargos(uu):
    cargos = []
    literalcargos = []
    literalcargos.append(uu.split("|")[0])
    literalcargos.append(uu.split("|")[1])
    for literal in literalcargos:
        previocargos = literal.split(";")
        for a in previocargos:
            if a[:1] == "[":
                intcargos = a[1:-1].split("-")
                for b in range(int(intcargos[0]), int(intcargos[1]) + 1, 1):
                    cargos.append(b)
            else:
                if a != "":
                    cargos.append(a)
    if len(cargos) == 1:
        devolucion = "CARGO "
    else:
        devolucion = "CARGOS "
    for cargo in cargos:
        if cargo != "":
            devolucion = devolucion + str(cargo) + ", "
    return devolucion[:-2]


def DevuelveCargos(uu):
    cargos = []
    literalcargos = []
    literalcargos.append(uu.split("|")[0])
    literalcargos.append(uu.split("|")[1])
    for literal in literalcargos:
        previocargos = literal.split(";")
        for a in previocargos:
            if a[:1] == "[":
                intcargos = a[1:-1].split("-")
                for b in range(int(intcargos[0]), int(intcargos[1]) + 1, 1):
                    cargos.append(b)
            else:
                if a != "":
                    cargos.append(a)
    return cargos


def descomponeUU(uu):
    a = ['', '']
    if uu >= 16:
        a[0] = 16
        a[1] = uu - 16
    elif uu < 16:
        a[0] = uu
        a[1] = 0
    return a


def exportar(form):
    ruta = QFileDialog.getOpenFileName(form, u"Seleccionar fichero de facturación", "", "Ficheros Excel (*.xls *.xlsx)")
    if ruta != "":
        wb = openpyxl.load_workbook(unicode(ruta))
        nhoja, ok = QInputDialog.getText(form, "Nombre de la nueva hoja", "Introduzca nombre:")
        hj = wb.create_sheet(str(nhoja))
        bg = PatternFill(fill_type='solid', start_color='C4D79B', end_color='C4D79B')
        ag = Alignment(horizontal='center', vertical='center')
        borde = Border(left=Side(border_style='thin', color='000000'),
                       right=Side(border_style='thin', color='000000'),
                       top=Side(border_style='thin', color='000000'),
                       bottom=Side(border_style='thin', color='000000'))
        fuente = Font(name='Calibri',
                      size=10,
                      bold=True,
                      color='000000')

        for c in range(form.tabla_2.columnCount()):
            d = hj.cell(row=1, column=c + 1, value=str(form.tabla_2.horizontalHeaderItem(c).text()))
            d.fill = bg
            d.alignment = ag
            d.border = borde
            d.font = fuente
        anchos = [('A', 21.42578125), ('B', 18.140625), ('C', 10.140625), ('D', 42.140625), ('E', 10.42578125),
                  ('F', 46.85546875), ('G', 13.0), ('H', 14.42578125), ('I', 14.85546875), ('J', 9.7109375), ('K', 23.28515625),
                  ('L', 9.5703125), ('M', 9.85546875), ('N', 9.85546875), ('O', 17.85546875), ('P', 13.7109375),
                  ('Q', 7.85546875), ('R', 13.0), ('S', 12.140625), ('T', 19.0), ('U', 16.42578125),
                  ('V', 14.0), ('W', 19.42578125), ('X', 16.140625), ('Y', 20.140625), ('Z', 13.85546875),
                  ('AA', 125.0)]

        for columna, ancho in anchos:
            hj.column_dimensions[columna].width = ancho
        borde = Border(left=Side(border_style='thin', color='00000000'),
                       right=Side(border_style='thin', color='00000000'),
                       top=Side(border_style='thin', color='00000000'),
                       bottom=Side(border_style='thin', color='00000000'))
        fuente = Font(name='Calibri', size=10)
        for fila in range(form.tabla_2.rowCount()):
            fondoverde = PatternFill(fill_type='solid', start_color='E2EFDA', end_color='E2EFDA')
            for columna in range(form.tabla_2.columnCount()):
                if form.tabla_2.item(fila, columna) is not None:
                    d = hj.cell(row=fila + 2, column=columna + 1, value=str(form.tabla_2.item(fila, columna).text()))

                else:
                    d = hj.cell(row=fila + 2, column=columna + 1, value="")
                d.border = borde
                d.font = fuente
                if form.tabla_2.item(fila, 0) is not None:
                    if form.tabla_2.item(fila, 0).text() == "3":
                        d.fill = fondoverde

        for col in hj.columns:
            if col[0].column in ['M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'Y', 'Z']:
                for cell in col:
                    try:
                        a = float(cell.value)
                        cell.value = a
                        cell.number_format = "#,##0.00"
                    except:
                        pass
        totalfilas = form.tabla_2.rowCount()
        borde = Border(left=Side(border_style='medium', color='00000000'),
                       right=Side(border_style='medium', color='00000000'),
                       top=Side(border_style='medium', color='00000000'),
                       bottom=Side(border_style='medium', color='00000000'))
        fuente = Font(name='Calibri',
                      size=10,
                      bold=True)
        bg = PatternFill(fill_type='solid', start_color='FFFF99', end_color='FFFF99')

        a = hj.cell(row=totalfilas + 3, column=16, value="=SUM(P2:P" + str(totalfilas + 1) + ")")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a.number_format = "#,##0.00"
        a = hj.cell(row=totalfilas + 3, column=15, value=u"TOTAL LIQUIDACIÓN")
        a.fill = bg
        a.font = fuente
        a.border = borde
        #a.number_format = "#,##0.00"
        a = hj.cell(row=totalfilas + 3, column=22, value="=SUM(V2:V" + str(totalfilas + 1) + ")")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a.number_format = "#,##0.00"
        a = hj.cell(row=totalfilas + 3, column=21, value=u"TOTAL INC PADRÓN")
        a.fill = bg
        a.font = fuente
        a.border = borde
        bg = PatternFill(fill_type='solid', start_color='FABF8F', end_color='FABF8F')
        a = hj.cell(row=totalfilas + 3, column=25, value="TOTAL SIN IVA")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 4, column=25, value="TOTAL CON IVA")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 5, column=25, value="IVA")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 6, column=25, value="UU")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 7, column=25, value="precio uu")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 8, column=25, value="UU=0")
        a.fill = bg
        a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 3, column=26, value="=SUM(Z2:Z" + str(totalfilas + 1) + ")")
        a.fill = bg
        #a.font = fuente
        a.number_format = "#,##0.00"
        a.border = borde
        a = hj.cell(row=totalfilas + 4, column=26, value="=SUM(Z" + str(totalfilas + 3) + "*1.21)")
        a.fill = bg
        #a.font = fuente
        a.number_format = "#,##0.00"
        a.border = borde
        a = hj.cell(row=totalfilas + 5, column=26, value="=Z" + str(totalfilas + 4) + "-Z" + str(totalfilas + 4))
        a.fill = bg
        #a.font = fuente
        a.number_format = "#,##0.00"
        a.border = borde
        a = hj.cell(row=totalfilas + 6, column=26, value=str(totalfilas))
        a.fill = bg
        #a.font = fuente
        a.border = borde
        a = hj.cell(row=totalfilas + 7, column=26, value="=Z" + str(totalfilas + 3) + "/Z" + str(totalfilas + 6))
        a.fill = bg
        #a.font = fuente
        a.border = borde
        a.number_format = "#,##0.00"
        a = hj.cell(row=totalfilas + 8, column=26, value="[...]")
        a.fill = bg
        #a.font = fuente
        a.border = borde


        try:
            wb.save(unicode(ruta))
            QMessageBox.information(form, "Guardado", u"Guardado correctamente")
        except IOError as e:
            if e.errno == 13:
                QMessageBox.critical(form, "Error", u"El fichero de destino está en uso")


def getTI(id_subo):
    path = os.getcwd() + "/profiles/tipos.accdb"

    conn_str = (
        r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
        r'DBQ=' + path + ';'
    )
    cnxn = pyodbc.connect(conn_str)
    crsr = cnxn.cursor()
    sql = "SELECT TIPO_IMPOSITIVO FROM TI WHERE ID_SUBO=" + str(id_subo)
    crsr.execute(sql)
    row = crsr.fetchone()
    if row:
        return str(row.TIPO_IMPOSITIVO).replace(",", ".")
    else:
        return "MANUAL"


def CargaValores(form):
    path = QFileDialog.getOpenFileName(form, u"Seleccionar Access de Explotación", "", "Ficheros Excel (*.mdb *.accdb)")
    path = unicode(path)
    p = form.loading("Importando valores desde Access", form.tabWidget)
    p.show()
    conn_str = (
        r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
        r'DBQ=' + path + ';'
    )
    cnxn = pyodbc.connect(conn_str)
    crsr = cnxn.cursor()
    sql = "SELECT temp.REFERENCIA_CATASTRAL AS RC, NIF_SP_VALO AS NIF, NOMBRE_SP_VALO AS SP, temp.IMPORTE AS IMP, " \
          "SWITCH(temp.PER = temp.MINPER, temp.MINPER," \
          "temp.PER <> temp.MINPER, temp.MINPER & '-' & temp.PER) AS PERI " \
          "FROM(" \
          "SELECT NOMBRE_SP_VALO, NIF_SP_VALO, REFERENCIA_CATASTRAL, " \
          "CINT(PERIODO) AS PER FROM [LISTADO 1] ) AS ini " \
          "INNER JOIN (" \
          "SELECT MAX(CINT(PERIODO))AS PER, SUM(IMPORTE_VALO) AS IMPORTE, " \
          "MIN(CINT(PERIODO)) AS MINPER, REFERENCIA_CATASTRAL FROM [LISTADO 1] " \
          "GROUP BY REFERENCIA_CATASTRAL) as temp " \
          "ON (ini.REFERENCIA_CATASTRAL = temp.REFERENCIA_CATASTRAL AND ini.PER= temp.PER);"

    crsr.execute(sql)
    rows = crsr.fetchall()
    for row in rows:
        RC = row.RC
        SP = row.SP
        NIF = row.NIF
        PERI = row.PERI
        IMP = row.IMP
        IMP = float(IMP) / 100
        for i in range(form.tabla_2.rowCount()):
            QApplication.processEvents()
            refcat = form.tabla_2.item(i, 5).text()
            if RC == refcat:
                form.tabla_2.setItem(i, 18, QTableWidgetItem(str(SP)))
                form.tabla_2.setItem(i, 17, QTableWidgetItem(str(NIF)))
                form.tabla_2.setItem(i, 21, QTableWidgetItem(str(IMP)))
                form.tabla_2.setItem(i, 16, QTableWidgetItem(str(PERI)))
    sql = "SELECT temp.REFERENCIA_CATASTRAL AS RC, NIF_SP_VALO AS NIF, NOMBRE_SP_VALO AS SP, temp.IMPORTE AS IMP, " \
          "SWITCH(temp.PER = temp.MINPER, temp.MINPER," \
          "temp.PER <> temp.MINPER, temp.MINPER & '-' & temp.PER) AS PERI " \
          "FROM(" \
          "SELECT NOMBRE_SP_VALO, NIF_SP_VALO, REFERENCIA_CATASTRAL, " \
          "CINT(PERIODO) AS PER FROM [LISTADO 2] ) AS ini " \
          "INNER JOIN (" \
          "SELECT MAX(CINT(PERIODO))AS PER, SUM(IMPORTE_VALO) AS IMPORTE, " \
          "MIN(CINT(PERIODO)) AS MINPER, REFERENCIA_CATASTRAL FROM [LISTADO 2] " \
          "GROUP BY REFERENCIA_CATASTRAL) as temp " \
          "ON (ini.REFERENCIA_CATASTRAL = temp.REFERENCIA_CATASTRAL AND ini.PER= temp.PER);"

    crsr.execute(sql)
    rows = crsr.fetchall()
    for row in rows:
        RC = row.RC
        SP = row.SP
        NIF = row.NIF
        PERI = row.PERI
        IMP = row.IMP
        IMP = float(IMP) / 100
        for i in range(form.tabla_2.rowCount()):
            QApplication.processEvents()
            refcat = form.tabla_2.item(i, 5).text()
            if RC == refcat:
                form.tabla_2.setItem(i, 18, QTableWidgetItem(str(SP)))
                form.tabla_2.setItem(i, 17, QTableWidgetItem(str(NIF)))
                form.tabla_2.setItem(i, 21, QTableWidgetItem(str(IMP)))
                form.tabla_2.setItem(i, 16, QTableWidgetItem(str(PERI)))
    del crsr
    cnxn.close()
    p.close()


def lc():
    print "a"
    anchos = []
    wb = openpyxl.load_workbook(unicode("D:/ahlopez/Escritorio/factalc.xlsx"))
    print wb
    ws = wb.get_sheet_by_name("1er SEMESTRE 2017")
    contador = 0
    for col in ws.columns:
        contador = contador + 1
        arf = (col[0].column, ws.column_dimensions[col[0].column].width)
        anchos.append(arf)
    print anchos


def jj():
    print "a"
    anchos = []
    wb = openpyxl.load_workbook(unicode("D:/ahlopez/Escritorio/fact - copia.xlsx"))
    print wb
    ws = wb.get_sheet_by_name("kilo")
    cell = ws.cell(row=147, column=30)
    print cell.number_format
