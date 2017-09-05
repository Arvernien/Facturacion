# -*- coding: utf-8 -*-
import ventana
from PyQt4.QtGui import *
from PyQt4.QtCore import *
import sys
import os
import openpyxl
import locale
from ti import Ui_frmTI
import pyodbc
from win32com.client import Dispatch
import pythoncom
from threading import Thread
from Queue import Queue

from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

class frmTI(QWidget, Ui_frmTI):
    def __init__(self):
        super(frmTI, self).__init__(parent=None)
        self.setupUi(self)
        self.tabla.setColumnCount(4)
        self.tabla.setHorizontalHeaderLabels(['PROVINCIA', 'ID SUBO', 'MUNICIPIO', 'TIPO'])
        self.getTI()
        self.criterio.textChanged.connect(self.filtra)

    def filtra(self):
        print self.tabla.rowCount()
        for i in range(self.tabla.rowCount()):
            if str(self.criterio.text()).lower() not in str(self.tabla.item(i, 2).text()).lower():
                self.tabla.setRowHidden(i, True)
            else:
                self.tabla.setRowHidden(i, False)

    def getTI(self):
        path = os.getcwd() + "/profiles/tipos.accdb"

        conn_str = (
            r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
            r'DBQ=' + path + ';'
        )
        cnxn = pyodbc.connect(conn_str)
        crsr = cnxn.cursor()
        sql = "SELECT * FROM TI"
        crsr.execute(sql)
        rows = crsr.fetchall()
        for row in rows:
            self.tabla.setRowCount(self.tabla.rowCount() + 1)
            self.tabla.setItem(self.tabla.rowCount() - 1, 0, QTableWidgetItem(unicode(row.PROVINCIA)))
            self.tabla.setItem(self.tabla.rowCount() - 1, 1, QTableWidgetItem(unicode(row.ID_SUBO)))
            self.tabla.setItem(self.tabla.rowCount() - 1, 2, QTableWidgetItem(unicode(row.NOMBRE_MUNICIPIO)))
            self.tabla.setItem(self.tabla.rowCount() - 1, 3, QTableWidgetItem(unicode(row.TIPO_IMPOSITIVO)))

class Check(QWidget):
    def __init__(self):
        super(Check, self).__init__(parent=None)
        self.lyt = QHBoxLayout()
        self.chk = QCheckBox()
        self.chk.setCheckState(Qt.Checked)
        self.lyt.setAlignment(Qt.AlignHCenter)
        self.lyt.addWidget(self.chk)
        self.lyt.setContentsMargins(0, 0, 0, 0)
        self.setLayout(self.lyt)

    def isChecked(self):
        return self.chk.isChecked()

class Principal(QMainWindow, ventana.Ui_MainWindow):
    class loading(QWidget):
        def __init__(self, texto, parent):
            QWidget.__init__(self, parent=None)
            self.movie = QMovie(":/img/loading.gif", QByteArray(), self)
            self.movie.setScaledSize(QSize(25, 25))
            size = self.movie.scaledSize()
            self.titulo = QLabel()
            self.titulo.setText(texto)
            self.titulo.setAlignment(Qt.AlignHCenter)
            self.titulo.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
            self.movie_screen = QLabel()
            self.movie_screen.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            self.movie_screen.setAlignment(Qt.AlignCenter)
            self.setWindowFlags(Qt.FramelessWindowHint)
            self.setAttribute(Qt.WA_TranslucentBackground)
            self.setObjectName("principal")
            self.setStyleSheet("QWidget {border: 2px rgb(143,212,0);}")
            center = parent.mapToGlobal(parent.frameGeometry().center())
            x = center.x() - 75
            y = center.y() - 25
            self.setGeometry(x, y, 150, 50)
            main_layout = QVBoxLayout()
            main_layout.addWidget(self.titulo)
            main_layout.addWidget(self.movie_screen)
            self.setLayout(main_layout)
            self.movie.setCacheMode(QMovie.CacheAll)
            self.movie.setSpeed(100)
            self.movie_screen.setMovie(self.movie)
            self.movie.start()
            self.marco = parent
            self.marco.setEnabled(False)

        def __del__(self):
            self.marco.setEnabled(True)

        def paintEvent(self, event):
            # get current window size
            s = self.size()
            qp = QPainter()
            qp.begin(self)
            qp.setRenderHint(QPainter.Antialiasing, True)
            qp.setPen(QColor(qRgb(143, 212, 0)))
            qp.setBrush(QColor(qRgb(255, 255, 255)))
            qp.drawRoundedRect(0, 0, s.width(), s.height(), 10, 10)
            qp.end()

    def __init__(self, parent=None):
        super(Principal, self).__init__(parent)
        self.setupUi(self)
        reload(sys)
        sys.setdefaultencoding('utf-8')
        self.loadProfiles()
        self.actionCargar_Excel_de_Facturaci_n.triggered.connect(self.cargaExcel)
        self.actionImportar_Fichero_de_carga.triggered.connect(self.ficherocarga)
        self.actionImportar_datos_desde_FIN_de_Retorno.triggered.connect(self.finRetorno)
        self.actionA_Excel_de_facturaci_n.triggered.connect(self.exporta)
        self.actionImportar_Access_Valores.triggered.connect(self.leevalores)
        self.but_TI.clicked.connect(self.abreTI)
        self.but_open.clicked.connect(self.cargaExcel)
        self.but_Calcular.clicked.connect(self.calcula)
        self.pushButton_3.clicked.connect(self.factura)
        self.but_Calcular.setEnabled(False)
        self.actionImportar_Access_Valores.setVisible(False)
        self.selectorPerfiles.currentIndexChanged.connect(self.cargaperfil)
        self.actionGenerar_Access_para_BET.triggered.connect(self.exportaBet)
        self.headers = self.tabla_2.horizontalHeader()
        self.headers.setContextMenuPolicy(Qt.CustomContextMenu)
        self.headers.customContextMenuRequested.connect(self.activamenu)
        self.headers.setSelectionMode(QAbstractItemView.SingleSelection)


    def abreTI(self):
        self.w = frmTI()
        self.w.show()
        return self.w

    def activamenu(self, position):
        menu = QMenu()
        iconrellenar = QIcon()
        iconrellenar.addPixmap(QPixmap(":/img/pencil.png"))
        iconborrar = QIcon()
        iconborrar.addPixmap(QPixmap(":/img/x"))
        rellenar = menu.addAction(iconrellenar, "Introducir dato")
        delete = menu.addAction(iconborrar, "Eliminar datos")
        ac = menu.exec_(self.tabla_2.mapToGlobal(position))
        column = self.headers.logicalIndexAt(position)
        ncol = self.tabla_2.horizontalHeaderItem(column).text()
        if ac == rellenar:
            dato, ok = QInputDialog.getText(self, "Datos para la columna " + ncol + ":", "Datos:                                                                                 ")
            if ok == True:
                self.rellena(column, dato)
            else:
                pass
        elif ac == delete:
            ok = QMessageBox.question(self, "Eliminar datos", "Se van a eliminar todos los datos de la columna " + ncol
                                      + u". ¿Está seguro?",
                                      QMessageBox.Yes | QMessageBox.No)
            if ok == QMessageBox.Yes:
                self.rellena(column, "")
            else:
                print "columna:" + str(column)

    def rellena(self, columna, dato):
        for i in range(self.tabla_2.rowCount()):
            self.tabla_2.setItem(i, columna, QTableWidgetItem(dato))



    def loadProfiles(self):
        for dir, subdir, files in os.walk('./profiles'):
            for file in files:
                if file.split(".")[0][:1] != "_" and file.split(".")[1] == "py":
                    self.selectorPerfiles.addItem(file.split(".")[0])

    def cargaExcel(self):
        ruta = QFileDialog.getOpenFileName(self, "Seleccione Excel exportado del control", "", "Ficheros Excel (*.xls *.xlsx)")
        if ruta != "":
            wb = openpyxl.load_workbook(unicode(ruta))
            hoja = wb.get_sheet_by_name('Hoja1')
            self.tabla.setRowCount(hoja.max_row-1)
            self.tabla.setColumnCount(hoja.max_column+1)
            cabeceras = ['']
            for i in range(1, hoja.max_column+1, 1):
                cabeceras.append(hoja.cell(row=1, column=i).value)
            self.tabla.setHorizontalHeaderLabels(cabeceras)
            self.tabla.horizontalHeader().setResizeMode(QHeaderView.ResizeToContents)
            for fila in range(1, hoja.max_row+1, 1):
                self.tabla.setCellWidget(fila-1, 0, Check())
                for col in range(1, hoja.max_column+1, 1):
                    self.tabla.setItem(fila-1, col, QTableWidgetItem(unicode(str(hoja.cell(row=fila+1, column=col).value))))

    def factura(self):
        try:
            self.tabla_2.setUpdatesEnabled(False)
            mod = __import__(str('profiles.' + self.selectorPerfiles.currentText()), fromlist=['*'])
            p = self.loading("Aplicando perfil seleccionado", self.tabWidget)
            p.show()
            mod.confTabla(self)

        finally:
            self.tabla_2.setUpdatesEnabled(True)
            if p:
                p.close()
            mod.activarFunciones(self)




    def calcula(self):
        try:
            self.tabla_2.setUpdatesEnabled(False)
            mod = __import__(str('profiles.' + self.selectorPerfiles.currentText()), fromlist=['*'])
            thread = Thread(target=mod.calcular, args=(self,))
            p = self.loading("Calculando", self.tabWidget)
            p.show()
            QApplication.processEvents()
            thread.start()
            while thread.isAlive():
                QApplication.processEvents()
            p.close()
        finally:
            self.tabla_2.setUpdatesEnabled(True)
            if p:
                p.close()

    def ficherocarga(self):
        mod = __import__(str('profiles.' + self.selectorPerfiles.currentText()), fromlist=['*'])
        mod.importaCarga(self)

    def finRetorno(self):
        try:
            self.tabla_2.setUpdatesEnabled(False)
            mod = __import__(str('profiles.' + self.selectorPerfiles.currentText()), fromlist=['*'])
            mod.importaFin(self)
        finally:
            self.tabla_2.setUpdatesEnabled(True)


    def desactivarFunciones(self):
        self.actionImportar_Fichero_de_carga.setEnabled(False)
        self.actionImportar_datos_desde_FIN_de_Retorno.setEnabled(False)
        self.but_Calcular.setEnabled(False)
        self.actionGenerar_Access_para_BET.setVisible(False)
        self.actionImportar_Access_Valores.setVisible(False)


    def exporta(self):
        mod = __import__(str('profiles.' + self.selectorPerfiles.currentText()), fromlist=['*'])
        mod.exportar(self)

    def cargaperfil(self):
        self.desactivarFunciones()
        mod = __import__(str('profiles.' + self.selectorPerfiles.currentText()), fromlist=['*'])
        mod.configura(self)

    def exportaBet(self):
        mod = __import__(str('profiles.' + self.selectorPerfiles.currentText()), fromlist=['*'])
        mod.GenBet(self)

    def leevalores(self):
        try:
            self.tabla_2.setUpdatesEnabled(False)
            mod = __import__(str('profiles.' + self.selectorPerfiles.currentText()), fromlist=['*'])
            mod.CargaValores(self)
        finally:
            self.tabla_2.setUpdatesEnabled(True)




def main(argv):
    locale.setlocale(locale.LC_TIME, "esp_esp")
    app = QApplication(argv)
    translator = QTranslator(app)
    j = QLocale.system().name()
    path = QLibraryInfo.location(QLibraryInfo.TranslationsPath)
    translator.load('qt_%s' % j, path)
    app.installTranslator(translator)
    global form
    form = Principal()
    form.show()
    app.exec_()
    return form


if __name__ == '__main__':
    main(sys.argv)
